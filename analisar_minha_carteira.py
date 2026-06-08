# -*- coding: utf-8 -*-
"""
analisar_minha_carteira.py
══════════════════════════════════════════════════════════════════════
Script de uso do leitor. Preencha o arquivo minha_carteira.xlsx com
seus tickers e pesos, aponte para seu CSV de preços e execute.

Uso:
    python analisar_minha_carteira.py

Saída:
    resultado_carteira.xlsx  (na mesma pasta)

══════════════════════════════════════════════════════════════════════
REQUISITOS:
    pip install pandas numpy scipy openpyxl

FORMATO DO CSV DE PREÇOS:
    - Primeira coluna: data (YYYY-MM-DD)
    - Demais colunas: tickers B3 (ex: PETR4, VALE3, IBOV, DIVO11...)
    - Pelo menos uma coluna chamada "IBOV" para o benchmark

DISCLAIMERS:
  • Análise retroativa. NÃO é recomendação de investimento.
  • O autor do script não é analista credenciado pela CVM.
  • Desempenho passado não garante desempenho futuro.
══════════════════════════════════════════════════════════════════════
"""

import os, math, warnings
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from scipy.optimize import minimize, Bounds

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════
# CONFIGURAÇÕES — edite aqui
# ══════════════════════════════════════════════════════════════════════
PASTA        = os.path.dirname(os.path.abspath(__file__))
CSV_PRECOS   = os.path.join(PASTA, "precos_b3.csv")         # seu CSV de preços
ENTRADA      = os.path.join(PASTA, "minha_carteira.xlsx")   # seu template preenchido
SAIDA        = os.path.join(PASTA, "resultado_carteira.xlsx")

RF_AA   = 0.1440   # taxa livre de risco anual (ajuste para a Selic atual)
ALPHA   = 0.5      # limite inferior = 50% do peso original
BETA    = 2.0      # limite superior = 200% do peso original

BENCHMARK_IBOV  = "IBOV"    # coluna do IBOV no CSV
BENCHMARK_DIVO  = "DIVO11"  # coluna do benchmark de dividendos (opcional)

# ══════════════════════════════════════════════════════════════════════
# HELPERS DE ESTILO
# ══════════════════════════════════════════════════════════════════════
def fill(h):   return PatternFill("solid", fgColor=h)
def fnt(b=False, c="1A1A2E", s=10, i=False):
    return Font(bold=b, color=c, size=s, italic=i, name="Calibri")
def aln(h="center", w=False): return Alignment(horizontal=h, vertical="center", wrap_text=w)
def brd(c="BFBFBF"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def cel(ws, r, c, v=None, **kw):
    cell = ws.cell(r, c)
    if not isinstance(cell, MergedCell):
        if v is not None: cell.value = v
        if "fl" in kw: cell.fill = kw["fl"]
        if "fn" in kw: cell.font = kw["fn"]
        if "al" in kw: cell.alignment = kw["al"]
        if "br" in kw: cell.border = kw["br"]
        if "nf" in kw: cell.number_format = kw["nf"]
    return cell

C1="1F3864"; C2="FFFFFF"; C3="FFF2CC"; C4="C6EFCE"; C5="FCE4D6"
C6="F5F5F5"; CO="C55A11"

# ══════════════════════════════════════════════════════════════════════
# CÁLCULO DE MÉTRICAS
# ══════════════════════════════════════════════════════════════════════
RF_D = math.log(1 + RF_AA) / 252

def metricas(w, lr, ibov_lr=None, divo_lr=None):
    lr    = np.nan_to_num(lr)
    mu_a  = np.mean(lr, 0) * 252
    cov_a = np.atleast_2d(np.cov(lr.T, ddof=1) * 252)
    Ra    = float(w @ mu_a)
    vol   = float(np.sqrt(np.clip(w @ cov_a @ w, 0, None)))
    sharpe = (Ra - RF_AA) / vol if vol > 1e-12 else -999.0

    pr     = lr @ w
    excess = pr - RF_D
    neg    = excess[excess < 0]
    sd_dn  = math.sqrt(np.mean(neg**2) * 252) if len(neg) else 1e-12
    sortino = (Ra - RF_AA) / sd_dn

    var95  = float(np.percentile(pr, 5))
    cvar95 = float(np.mean(pr[pr <= var95]))
    cum    = np.exp(np.cumsum(pr))
    peak   = np.maximum.accumulate(cum)
    mdd    = float(np.min((cum - peak) / peak))
    cum_r  = float(cum[-1] - 1)

    beta = alpha_j = treynor = ir_divo = None
    if ibov_lr is not None:
        ib    = np.nan_to_num(ibov_lr[:len(pr)])
        var_ib = float(np.var(ib, ddof=1))
        cov_pi = float(np.cov(pr, ib, ddof=1)[0, 1])
        beta   = cov_pi / var_ib if var_ib > 1e-14 else 1.0
        Ra_ib  = float(np.mean(ib) * 252)
        alpha_j = Ra - (RF_AA + beta * (Ra_ib - RF_AA))
        treynor = (Ra - RF_AA) / beta if abs(beta) > 1e-12 else 0.0
    if divo_lr is not None:
        dv     = np.nan_to_num(divo_lr[:len(pr)])
        active = pr - dv
        te     = float(np.std(active, ddof=1) * math.sqrt(252))
        ir_divo = float(np.mean(active) * 252 / te) if te > 1e-12 else 0.0

    return dict(Ra=Ra, vol=vol, sharpe=sharpe, sortino=sortino,
                var95=var95, cvar95=cvar95, mdd=mdd, cum_ret=cum_r,
                beta=beta, alpha_j=alpha_j, treynor=treynor, ir_divo=ir_divo)

def otimizar(w_rec, lr):
    lr  = np.nan_to_num(lr)
    EPS = 1e-4
    lb  = np.maximum(ALPHA * w_rec, EPS)
    ub  = np.minimum(BETA  * w_rec, 1.0)
    if np.sum(lb) > 1.0 + 1e-6: lb = lb / np.sum(lb) * 0.99
    ub  = np.maximum(ub, lb + 1e-5)
    mu_a  = np.mean(lr, 0) * 252
    cov_a = np.atleast_2d(np.cov(lr.T, ddof=1) * 252)
    def neg_sh(w):
        r = float(w @ mu_a)
        v = float(np.sqrt(np.clip(w @ cov_a @ w, 0, None)))
        return -(r - RF_AA) / v if v > 1e-12 else 999.0
    w0  = np.clip(w_rec.copy(), lb, ub); w0 /= w0.sum()
    res = minimize(neg_sh, w0, method="SLSQP",
                   bounds=Bounds(lb=lb, ub=ub),
                   constraints=[{"type": "eq", "fun": lambda w: w.sum() - 1.0}],
                   options={"maxiter": 5000, "ftol": 1e-14})
    w_opt = np.clip(res.x, lb, ub); w_opt /= w_opt.sum()
    return w_opt, res.success

# ══════════════════════════════════════════════════════════════════════
# LEITURA DO TEMPLATE
# ══════════════════════════════════════════════════════════════════════
print("═" * 60)
print("  Analisador de Carteira Personalizada")
print("  github.com/alissondpoliveira/Markowitz-carteiras-dividendos")
print("═" * 60)

if not os.path.exists(ENTRADA):
    print(f"\n[ERRO] Arquivo não encontrado: {ENTRADA}")
    print("  Certifique-se de que minha_carteira.xlsx está na mesma pasta.")
    exit(1)

if not os.path.exists(CSV_PRECOS):
    print(f"\n[ERRO] CSV de preços não encontrado: {CSV_PRECOS}")
    print("  Veja o README.md para instruções sobre o formato esperado.")
    exit(1)

wb_in = openpyxl.load_workbook(ENTRADA)
ws_in = wb_in["Minha Carteira"]

portfolio = {}
nome_carteira = "Minha Carteira"

for row in ws_in.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1] is not None:
        ticker = str(row[0]).strip().upper()
        peso   = float(row[1])
        if ticker.lower() == "nome_carteira":
            nome_carteira = str(row[1])
            continue
        if peso > 0:
            portfolio[ticker] = peso

if not portfolio:
    print("[ERRO] Nenhum ativo encontrado em minha_carteira.xlsx")
    exit(1)

# Normalizar pesos
total = sum(portfolio.values())
portfolio = {k: v/total for k, v in portfolio.items()}

print(f"\nCarteira: {nome_carteira}")
print(f"Ativos: {len(portfolio)}")
for tk, w in portfolio.items():
    print(f"  {tk}: {w:.1%}")

# ══════════════════════════════════════════════════════════════════════
# DADOS DE PREÇOS
# ══════════════════════════════════════════════════════════════════════
df_all  = pd.read_csv(CSV_PRECOS, index_col=0, parse_dates=True).sort_index().dropna(how="all")
colunas = df_all.columns.tolist()

# Verificar tickers
ausentes = [tk for tk in portfolio if tk not in colunas]
if ausentes:
    print(f"\n[AVISO] Tickers não encontrados no CSV e serão ignorados: {ausentes}")
    for tk in ausentes:
        del portfolio[tk]
    if not portfolio:
        print("[ERRO] Nenhum ativo válido após remoção dos ausentes.")
        exit(1)
    total = sum(portfolio.values()); portfolio = {k: v/total for k, v in portfolio.items()}
    print(f"  Pesos renormalizados para {len(portfolio)} ativos.")

tickers = list(portfolio.keys())
w_rec   = np.array(list(portfolio.values()), dtype=float)
w_rec  /= w_rec.sum()

# Log-retornos
df_port = df_all[tickers].ffill(limit=5)
arr     = df_port.values.astype(float)
lr_port = np.where(np.isnan(arr[1:]) | np.isnan(arr[:-1]), 0.0,
                   np.log(np.clip(arr[1:] / arr[:-1], 1e-8, None)))

def lr_serie(col):
    if col not in df_all.columns: return None
    s = df_all[col].ffill(limit=5).values.astype(float)
    return np.where(np.isnan(s[1:]) | np.isnan(s[:-1]), 0.0,
                    np.log(np.clip(s[1:] / s[:-1], 1e-8, None)))

ibov_lr = lr_serie(BENCHMARK_IBOV)
divo_lr = lr_serie(BENCHMARK_DIVO)

# ══════════════════════════════════════════════════════════════════════
# ANÁLISE
# ══════════════════════════════════════════════════════════════════════
print("\nCalculando métricas...")
m_rec = metricas(w_rec, lr_port, ibov_lr, divo_lr)

print("Otimizando pesos (Max Sharpe)...")
w_opt, ok = otimizar(w_rec, lr_port)
m_opt = metricas(w_opt, lr_port, ibov_lr, divo_lr)

m_ibov = metricas(np.array([1.0]), ibov_lr.reshape(-1,1), ibov_lr, divo_lr) if ibov_lr is not None else None
m_divo = metricas(np.array([1.0]), divo_lr.reshape(-1,1), ibov_lr, divo_lr) if divo_lr is not None else None

inicio = df_all.index[0].date(); fim = df_all.index[-1].date()
n_dias = len(lr_port)

# ══════════════════════════════════════════════════════════════════════
# EXCEL DE SAÍDA
# ══════════════════════════════════════════════════════════════════════
print(f"Gerando {SAIDA}...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resultado"
ws.column_dimensions["A"].width = 32
for c in ["B","C","D","E"]:
    ws.column_dimensions[c].width = 16

R = 1
# Cabeçalho
ws.merge_cells(f"A{R}:E{R}")
cel(ws, R, 1, f"Análise de Carteira: {nome_carteira}",
    fl=fill(C1), fn=fnt(b=True, c=C2, s=14), al=aln("left"), br=brd())
ws.row_dimensions[R].height = 26; R += 1

ws.merge_cells(f"A{R}:E{R}")
cel(ws, R, 1, f"Período: {inicio} → {fim}  ({n_dias} pregões)  |  RF = {RF_AA:.2%} a.a.  |  α={ALPHA}  β={BETA}  |  Otimizador: Max Sharpe (SLSQP)",
    fl=fill(C3), fn=fnt(c=CO, s=9, i=True), al=aln("left"), br=brd())
ws.row_dimensions[R].height = 18; R += 1

ws.merge_cells(f"A{R}:E{R}")
cel(ws, R, 1, "⚠ Análise retroativa. NÃO é recomendação de investimento. Rentabilidade passada não garante resultados futuros.",
    fl=fill(C5), fn=fnt(c="C00000", s=9, i=True), al=aln("left", w=True), br=brd())
ws.row_dimensions[R].height = 18; R += 2

# Cabeçalho colunas
for ci, (col, lbl) in enumerate(zip(range(1,6), ["Métrica","Rec. (orig.)","Otimizado","IBOV","DIVO11"])):
    cor = [C1,"505050","17375E","404040","7030A0"][ci]
    cel(ws, R, col, lbl, fl=fill(cor), fn=fnt(b=True, c=C2, s=10), al=aln("center"), br=brd())
R += 1

LINHAS = [
    ("─── RETORNO E RISCO ───",    None,None,None,None,"sep"),
    ("Retorno Anual",               "Ra","0.00%","0.00%","0.00%"),
    ("Volatilidade Anual",          "vol","0.00%","0.00%","0.00%"),
    ("Retorno Acumulado",           "cum_ret","0.00%","0.00%","0.00%"),
    ("─── INDICADORES RISCO-RETORNO ───", None,None,None,None,"sep"),
    ("Índice de Sharpe",            "sharpe","+0.00","+0.00","+0.00"),
    ("Índice de Sortino",           "sortino","+0.00","+0.00","+0.00"),
    ("Índice de Treynor",           "treynor","+0.000","+0.000","+0.000"),
    ("─── RISCO ───",              None,None,None,None,"sep"),
    ("Beta vs IBOV",                "beta","0.00","0.00","0.00"),
    ("Alpha de Jensen vs IBOV",     "alpha_j","0.00%","0.00%","0.00%"),
    ("VaR 95% (diário)",            "var95","0.000%","0.000%","0.000%"),
    ("CVaR 95% (diário)",           "cvar95","0.000%","0.000%","0.000%"),
    ("Maximum Drawdown",            "mdd","0.00%","0.00%","0.00%"),
    ("─── BENCHMARK DIVO11 ───",   None,None,None,None,"sep"),
    ("IR vs DIVO11",                "ir_divo","+0.00","+0.00","+0.00"),
]

for midx, linha in enumerate(LINHAS):
    if linha[5] == "sep" if len(linha) > 5 else False:
        ws.merge_cells(f"A{R}:E{R}")
        cel(ws, R, 1, linha[0], fl=fill("1E6B6B"), fn=fnt(b=True, c=C2, s=9), al=aln("center"), br=brd())
        ws.row_dimensions[R].height = 15; R += 1; continue

    label, key, nf_rec, nf_opt, nf_bm = linha[0], linha[1], linha[2], linha[3], linha[4]
    bg = C6 if midx % 2 == 0 else "FFFFFF"
    cel(ws, R, 1, label, fl=fill(bg), fn=fnt(b=True, s=10), al=aln("left"), br=brd())

    vals = [
        m_rec.get(key) if key else None,
        m_opt.get(key) if key else None,
        m_ibov.get(key) if (key and m_ibov) else None,
        m_divo.get(key) if (key and m_divo) else None,
    ]
    nfs  = [nf_rec, nf_opt, nf_bm, nf_bm]

    # Determinar melhor valor (para highlight)
    high_k = {"Ra","sharpe","sortino","treynor","cum_ret","ir_divo","alpha_j"}
    clean  = [v for v in vals[:2] if v is not None]  # só compara rec vs opt
    bval   = (max(clean) if key in high_k else min(clean, key=abs)) if clean else None

    for ci, (v, nf) in enumerate(zip(vals, nfs)):
        col = 2 + ci
        cel(ws, R, col, v, nf=nf if v is not None else "@",
            fl=fill(C4 if (bval is not None and v is not None and ci < 2 and abs(v - bval) < 1e-9) else bg),
            fn=fnt(s=10, c="404040" if ci >= 2 else "1A1A2E"),
            al=aln("center"), br=brd())
    R += 1

# Aba de pesos
ws2 = wb.create_sheet("Pesos")
ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 14

R2 = 1
ws2.merge_cells(f"A{R2}:D{R2}")
cel(ws2,R2,1,"Pesos: Recomendado vs. Otimizado",fl=fill(C1),fn=fnt(b=True,c=C2,s=12),al=aln("center"),br=brd())
ws2.row_dimensions[R2].height=22; R2+=2

for ci,(h,c) in enumerate(zip(["Ativo","Peso Rec.","Peso Otim.","Δ"],range(1,5))):
    cel(ws2,R2,c,h,fl=fill("2E75B6"),fn=fnt(b=True,c=C2,s=10),al=aln("center"),br=brd())
R2+=1

for i, (tk, wr) in enumerate(zip(tickers, w_rec)):
    wo = float(w_opt[i]); delta = wo - wr
    bg = C6 if i%2==0 else "FFFFFF"
    cel(ws2,R2,1,tk,fl=fill(bg),fn=fnt(b=True,s=10),al=aln("left"),br=brd())
    cel(ws2,R2,2,wr,fl=fill(bg),fn=fnt(s=10),al=aln("center"),br=brd(),nf="0.00%")
    cel(ws2,R2,3,wo,fl=fill(bg),fn=fnt(s=10),al=aln("center"),br=brd(),nf="0.00%")
    cel(ws2,R2,4,delta,
        fl=fill(C4 if delta>0.001 else (C5 if delta<-0.001 else bg)),
        fn=fnt(s=10),al=aln("center"),br=brd(),nf="+0.00%")
    R2+=1

# Aba disclaimers
ws3 = wb.create_sheet("Disclaimers")
ws3.column_dimensions["A"].width = 80
disclaimers = [
    "DISCLAIMERS E LIMITAÇÕES",
    "",
    "1. Esta análise é exclusivamente retroativa (backtesting). Os resultados refletem o comportamento passado",
    "   dos ativos no período analisado e NÃO constituem previsão de desempenho futuro.",
    "",
    "2. NÃO é recomendação de investimento, análise de valores mobiliários, nem consultoria financeira.",
    "   O script foi desenvolvido para fins educacionais e de pesquisa em finanças quantitativas.",
    "",
    "3. O otimizador maximiza o Índice de Sharpe histórico. Em janelas curtas de dados, o modelo",
    "   tende a superestimar o Sharpe in-sample (estimation error maximization). Resultados fora",
    "   da amostra são tipicamente inferiores aos resultados no período de estimação.",
    "",
    "4. Os indicadores (Sharpe, Sortino, Beta, VaR, CVaR, MDD) são calculados sobre log-retornos",
    "   de preços e NÃO incorporam dividendos, custos de transação, imposto de renda, spreads",
    "   bid-ask ou outros custos operacionais.",
    "",
    "5. O DIVO11 é utilizado como proxy do IDIV (Índice de Dividendos B3). Não é idêntico ao índice:",
    "   há come-cotas, taxa de administração e tracking error.",
    "",
    "6. Antes de tomar qualquer decisão de investimento, consulte um analista credenciado pela CVM",
    "   ou um consultor de investimentos registrado.",
    "",
    "Código-fonte e metodologia: github.com/alissondpoliveira/Markowitz-carteiras-dividendos",
]
for i, line in enumerate(disclaimers):
    ws3.row_dimensions[i+1].height = 16
    cel(ws3, i+1, 1, line,
        fl=fill(C1 if i==0 else ("F8F8F8" if i%2==0 else "FFFFFF")),
        fn=fnt(b=(i==0), c=(C2 if i==0 else "1A1A2E"), s=(11 if i==0 else 9)),
        al=aln("left"))

wb.save(SAIDA)
print(f"\n✓ Resultado salvo em: {SAIDA}")
print(f"\n{'─'*60}")
print(f"  Retorno Anual  (rec. / otim.): {m_rec['Ra']:.1%} / {m_opt['Ra']:.1%}")
print(f"  Sharpe         (rec. / otim.): {m_rec['sharpe']:.3f} / {m_opt['sharpe']:.3f}")
print(f"  Vol. Anual     (rec. / otim.): {m_rec['vol']:.1%} / {m_opt['vol']:.1%}")
print(f"  MDD            (rec. / otim.): {m_rec['mdd']:.1%} / {m_opt['mdd']:.1%}")
if ibov_lr is not None:
    print(f"  Beta vs IBOV   (rec. / otim.): {m_rec['beta']:.3f} / {m_opt['beta']:.3f}")
if divo_lr is not None:
    print(f"  IR vs DIVO11   (rec. / otim.): {m_rec['ir_divo']:.3f} / {m_opt['ir_divo']:.3f}")
print(f"\n⚠ Retroativo. Não é recomendação de investimento.")

# -*- coding: utf-8 -*-
"""
backtest_benchmark.py
══════════════════════════════════════════════════════════════════════
Complemento à análise principal das carteiras de dividendos (jun 2026).

Módulo 1 — BACKTEST OUT-OF-SAMPLE
  - Split 50/50: treina na primeira metade, avalia na segunda
  - Otimiza Max Sharpe na janela de treino com bounds α/β
  - Reporta métricas de treino vs. teste, degradação do Sharpe

Módulo 2 — BENCHMARK DIVO11 (proxy IDIV)
  - Compara cada carteira vs. DIVO11 no período completo e no teste
  - Information Ratio, Alpha, tracking error, alfa de Jensen

══════════════════════════════════════════════════════════════════════
DISCLAIMERS:
  • Análise retroativa — NÃO é recomendação de investimento.
  • Backtest 50/50 é apenas ilustrativo; split único não valida robustez.
  • DIVO11 é proxy do IDIV mas NÃO é idêntico (come-cotas, tracking error).
  • Desempenho passado não é garantia de desempenho futuro.
══════════════════════════════════════════════════════════════════════
"""

import os, math, numpy as np, pandas as pd, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from scipy.optimize import minimize, Bounds

# ══════════════════════════════════════════════════════════════════════
# PARÂMETROS
# ══════════════════════════════════════════════════════════════════════
PASTA    = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(PASTA, "precos_b3.csv")
OUT_PATH = os.path.join(PASTA, "backtest_benchmark.xlsx")

RF_AA     = 0.1440
RF_DIARIO = math.log(1 + RF_AA) / 252
ALPHA = 0.5
BETA  = 2.0

# ══════════════════════════════════════════════════════════════════════
# CARTEIRAS (pesos recomendados — idênticos ao analisar_carteiras.py)
# ══════════════════════════════════════════════════════════════════════
CARTEIRAS = {
    "XP Dividendos": {
        "PETR4":0.100,"PRIO3":0.050,"VALE3":0.125,"CPLE3":0.150,
        "ENGI11":0.050,"AXIA3":0.100,"VIVT3":0.050,"ALOS3":0.075,
        "B3SA3":0.100,"CXSE3":0.050,"ITUB4":0.150,
    },
    "BTG Dividendos": {
        "PETR4":0.10,"ITUB4":0.10,"VALE3":0.05,"BBDC4":0.10,
        "AXIA3":0.10,"EQTL3":0.10,"CPLE3":0.10,"CXSE3":0.10,
        "MOTV3":0.05,"CSMG3":0.05,"ALOS3":0.05,"CURY3":0.10,
    },
    "Itau BBA Div.": {
        "AXIA3":0.20,"ALOS3":0.20,"BBDC4":0.20,"VALE3":0.20,"PETR4":0.20,
    },
    "Santander Div.": {
        "ALOS3":0.10,"AXIA3":0.10,"BPAC11":0.10,"CPLE3":0.10,
        "VIVT3":0.10,"VALE3":0.10,"VBBR3":0.10,"CURY3":0.10,
        "ITUB4":0.10,"PETR3":0.10,
    },
    "BB Dividendos": {
        "ALOS3":0.10,"ABEV3":0.10,"BBDC4":0.10,"BRAP4":0.10,
        "CXSE3":0.10,"DIRR3":0.10,"ITSA4":0.10,"PETR4":0.10,
        "TAEE11":0.10,"TIMS3":0.10,
    },
}

CORES = {
    "XP Dividendos":  "C55A11",
    "BTG Dividendos": "17375E",
    "Itau BBA Div.":  "1E5799",
    "Santander Div.": "C00000",
    "BB Dividendos":  "375623",
    "DIVO11":         "7030A0",
    "IBOV":           "404040",
}
C_NAVY="1F3864"; C_WHITE="FFFFFF"; C_GRAY="F5F5F5"; C_YELLOW="FFF2CC"
C_GREEN="C6EFCE"; C_RED="FCE4D6"; C_BLUE="D6E4F0"; C_TEAL="1E6B6B"
C_ORANGE="C55A11"

def fill(h): return PatternFill("solid",fgColor=h)
def font(bold=False,color="1A1A2E",size=10,italic=False):
    return Font(bold=bold,color=color,size=size,italic=italic,name="Calibri")
def align(h="center",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def thin(c="BFBFBF"):
    s=Side(style="thin",color=c); return Border(left=s,right=s,top=s,bottom=s)

def sv(ws,r,c,v):
    cell=ws.cell(r,c)
    if not isinstance(cell,MergedCell): cell.value=v
    return cell

def fmt(ws,r,c,**kw):
    cell=ws.cell(r,c)
    if isinstance(cell,MergedCell): return
    if "fill"   in kw: cell.fill          = kw["fill"]
    if "font"   in kw: cell.font          = kw["font"]
    if "align"  in kw: cell.alignment     = kw["align"]
    if "border" in kw: cell.border        = kw["border"]
    if "nf"     in kw: cell.number_format = kw["nf"]

# ══════════════════════════════════════════════════════════════════════
# FUNÇÕES CORE
# ══════════════════════════════════════════════════════════════════════
def metricas(w, lr, ibov_lr=None, divo_lr=None):
    """Calcula métricas completas para vetor de pesos w."""
    lr = np.nan_to_num(lr)
    mu_a  = np.mean(lr, 0) * 252
    cov_raw = np.cov(lr.T, ddof=1) * 252
    cov_a = np.atleast_2d(cov_raw)  # garante 2D quando lr tem 1 coluna
    Ra  = float(w @ mu_a)
    vol = float(np.sqrt(np.clip(w @ cov_a @ w, 0, None)))
    sharpe = (Ra - RF_AA) / vol if vol > 1e-12 else -999.0

    pr = lr @ w
    excess = pr - RF_DIARIO
    neg    = excess[excess < 0]
    sd_dn  = math.sqrt(np.mean(neg**2) * 252) if len(neg) else 1e-12
    sortino = (Ra - RF_AA) / sd_dn

    var95 = float(np.percentile(pr, 5))
    cvar95 = float(np.mean(pr[pr <= var95]))
    cum    = np.exp(np.cumsum(pr))
    peak   = np.maximum.accumulate(cum)
    mdd    = float(np.min((cum - peak) / peak))
    cum_ret = float(cum[-1] - 1)  # retorno acumulado simples

    # Beta e alfa vs benchmark
    beta_ibov = alpha_ibov = treynor = ir_divo = None
    if ibov_lr is not None:
        ib = np.nan_to_num(ibov_lr[:len(pr)])
        var_ib = float(np.var(ib, ddof=1))
        cov_pi = float(np.cov(pr, ib, ddof=1)[0, 1])
        beta_ibov  = cov_pi / var_ib if var_ib > 1e-14 else 1.0
        Ra_ib      = float(np.mean(ib) * 252)
        alpha_ibov = Ra - (RF_AA + beta_ibov * (Ra_ib - RF_AA))  # alfa de Jensen
        treynor    = (Ra - RF_AA) / beta_ibov if abs(beta_ibov) > 1e-12 else 0.0

    if divo_lr is not None:
        dv = np.nan_to_num(divo_lr[:len(pr)])
        active = pr - dv
        te     = float(np.std(active, ddof=1) * math.sqrt(252))
        ir_val = float(np.mean(active) * 252 / te) if te > 1e-12 else 0.0
        ir_divo = ir_val

    return {
        "Ra": Ra, "vol": vol, "sharpe": sharpe, "sortino": sortino,
        "var95": var95, "cvar95": cvar95, "mdd": mdd, "cum_ret": cum_ret,
        "beta": beta_ibov, "alpha_j": alpha_ibov, "treynor": treynor,
        "ir_divo": ir_divo,
    }


def otimizar(w_rec, lr, ibov_lr=None):
    """Max Sharpe com bounds relativos α·wᵣₑc ≤ wᵢ ≤ β·wᵣₑc."""
    lr = np.nan_to_num(lr)
    n  = len(w_rec)
    EPS = 1e-4
    lb = np.maximum(ALPHA * w_rec, EPS)
    ub = np.minimum(BETA  * w_rec, 1.0)
    if np.sum(lb) > 1.0 + 1e-6:
        lb = lb / np.sum(lb) * 0.99
    ub = np.maximum(ub, lb + 1e-5)

    mu_a  = np.mean(lr, 0) * 252
    cov_raw = np.cov(lr.T, ddof=1) * 252
    cov_a = np.atleast_2d(cov_raw)  # garante 2D quando lr tem 1 coluna

    def neg_sharpe(w):
        r = float(w @ mu_a)
        v = float(np.sqrt(np.clip(w @ cov_a @ w, 0, None)))
        return -(r - RF_AA) / v if v > 1e-12 else 999.0

    w0 = np.clip(w_rec.copy(), lb, ub); w0 /= w0.sum()
    res = minimize(neg_sharpe, w0, method="SLSQP",
                   bounds=Bounds(lb=lb, ub=ub),
                   constraints=[{"type": "eq", "fun": lambda w: w.sum() - 1.0}],
                   options={"maxiter": 5000, "ftol": 1e-14})
    w_opt = np.clip(res.x, lb, ub); w_opt /= w_opt.sum()
    return w_opt, res.success


# ══════════════════════════════════════════════════════════════════════
# 1. CARREGAR DADOS
# ══════════════════════════════════════════════════════════════════════
print("─" * 65)
print("BACKTEST OUT-OF-SAMPLE + BENCHMARK DIVO11 — Junho 2026")
print("─" * 65)

df_all = pd.read_csv(CSV_PATH, index_col=0, parse_dates=True).sort_index()
df_all = df_all.dropna(how="all")
N_TOTAL = len(df_all)
SPLIT   = N_TOTAL // 2   # 50/50
df_train = df_all.iloc[:SPLIT]
df_test  = df_all.iloc[SPLIT:]

print(f"\nPeríodo completo : {df_all.index[0].date()} → {df_all.index[-1].date()} ({N_TOTAL} pregões)")
print(f"Treino (in-sample): {df_train.index[0].date()} → {df_train.index[-1].date()} ({len(df_train)} pregões)")
print(f"Teste  (out-of-sample): {df_test.index[0].date()} → {df_test.index[-1].date()} ({len(df_test)} pregões)")

# Benchmark DIVO11 e IBOV
def log_rets_serie(serie, index):
    s = serie.reindex(index).ffill(limit=5).values.astype(float)
    return np.where(np.isnan(s[1:]) | np.isnan(s[:-1]), 0.0, np.log(s[1:] / s[:-1]))

ibov_full  = log_rets_serie(df_all["IBOV"],   df_all.index)
ibov_train = log_rets_serie(df_train["IBOV"], df_train.index)
ibov_test  = log_rets_serie(df_test["IBOV"],  df_test.index)

divo_full  = log_rets_serie(df_all["DIVO11"],   df_all.index)
divo_train = log_rets_serie(df_train["DIVO11"], df_train.index)
divo_test  = log_rets_serie(df_test["DIVO11"],  df_test.index)

# Métricas do DIVO11
m_divo_full  = metricas(np.array([1.0]), divo_full.reshape(-1,1),  ibov_full,  divo_full)
m_divo_test  = metricas(np.array([1.0]), divo_test.reshape(-1,1),  ibov_test,  divo_test)

m_ibov_full  = metricas(np.array([1.0]), ibov_full.reshape(-1,1),  ibov_full,  divo_full)
m_ibov_test  = metricas(np.array([1.0]), ibov_test.reshape(-1,1),  ibov_test,  divo_test)

# ══════════════════════════════════════════════════════════════════════
# 2. PROCESSAR CADA CARTEIRA
# ══════════════════════════════════════════════════════════════════════
RESULTADOS = {}
print("\nProcessando carteiras...")

for nome, pesos in CARTEIRAS.items():
    tks = list(pesos.keys())
    w_rec = np.array(list(pesos.values()), dtype=float); w_rec /= w_rec.sum()

    # Preparar dados por janela
    def _prep(df_janela):
        df = df_janela[tks].ffill(limit=5).dropna(how="all")
        lr = np.nan_to_num(np.log(df.values[1:] / df.values[:-1]))
        ib = log_rets_serie(df_janela["IBOV"],   df_janela.index)[:len(lr)]
        dv = log_rets_serie(df_janela["DIVO11"], df_janela.index)[:len(lr)]
        return lr, ib, dv

    lr_full,  ib_full,  dv_full  = _prep(df_all)
    lr_train, ib_train, dv_train = _prep(df_train)
    lr_test,  ib_test,  dv_test  = _prep(df_test)

    # Otimização IN-SAMPLE (treino)
    w_opt_train, ok_train = otimizar(w_rec, lr_train, ib_train)

    # Métricas full (como recomendado)
    m_rec_full  = metricas(w_rec,        lr_full,  ib_full,  dv_full)
    # Métricas treino
    m_rec_train = metricas(w_rec,        lr_train, ib_train, dv_train)
    m_opt_train = metricas(w_opt_train,  lr_train, ib_train, dv_train)
    # Métricas TESTE (out-of-sample) — peso recomendado e peso otimizado no treino
    m_rec_test  = metricas(w_rec,        lr_test,  ib_test,  dv_test)
    m_opt_test  = metricas(w_opt_train,  lr_test,  ib_test,  dv_test)

    degradacao_sharpe = m_opt_test["sharpe"] - m_opt_train["sharpe"]

    RESULTADOS[nome] = {
        "tks": tks, "w_rec": w_rec, "w_opt_train": w_opt_train,
        "m_rec_full": m_rec_full,
        "m_rec_train": m_rec_train, "m_opt_train": m_opt_train,
        "m_rec_test":  m_rec_test,  "m_opt_test":  m_opt_test,
        "degradacao":  degradacao_sharpe, "ok": ok_train,
    }
    print(f"  {nome:<22}  Sharpe treino: {m_opt_train['sharpe']:.3f} → teste: {m_opt_test['sharpe']:.3f}  "
          f"(Δ {degradacao_sharpe:+.3f})  Rec. teste: {m_rec_test['sharpe']:.3f}")

# ══════════════════════════════════════════════════════════════════════
# 3. GERAR EXCEL
# ══════════════════════════════════════════════════════════════════════
print(f"\nGerando Excel: {OUT_PATH}")
wb = openpyxl.Workbook()

nomes = list(RESULTADOS.keys())

# ──────────────────────────────────────────────────────────────────
# ABA 1: BACKTEST OUT-OF-SAMPLE
# ──────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Backtest OOS"
ws.column_dimensions["A"].width = 28
for i in range(1, len(nomes)*3+4):
    ws.column_dimensions[get_column_letter(i+1)].width = 12

R = 1
sv(ws,R,1,"BACKTEST OUT-OF-SAMPLE · Carteiras de Dividendos · Junho 2026")
ws.merge_cells(start_row=R,start_column=1,end_row=R,end_column=len(nomes)*2+3)
ws.row_dimensions[R].height = 22
fmt(ws,R,1,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE,size=12),align=align("center"),border=thin())
R+=1

disc = (f"Split 50/50: Treino {df_train.index[0].date()} → {df_train.index[-1].date()} "
        f"({len(df_train)} pregões)  |  Teste {df_test.index[0].date()} → {df_test.index[-1].date()} "
        f"({len(df_test)} pregões)  |  RF = {RF_AA:.2%} a.a.  |  "
        f"α = {ALPHA}  β = {BETA}  |  Otimização: Max Sharpe in-sample, avaliada out-of-sample")
sv(ws,R,1,disc)
ws.merge_cells(start_row=R,start_column=1,end_row=R,end_column=len(nomes)*2+3)
ws.row_dimensions[R].height = 28
fmt(ws,R,1,fill=fill(C_YELLOW),font=font(size=9,italic=True,color=C_ORANGE,bold=True),
    align=align("center",wrap=True),border=thin())
R+=1

warn = ("⚠ LIMITAÇÃO: Split único 50/50 não valida robustez. Sharpe in-sample é tendencioso para cima (maximização de erro de estimação). "
        "O Sharpe out-of-sample é a métrica relevante. Degradação esperada — se for nula, suspeite de overfitting.")
sv(ws,R,1,warn)
ws.merge_cells(start_row=R,start_column=1,end_row=R,end_column=len(nomes)*2+3)
ws.row_dimensions[R].height = 28
fmt(ws,R,1,fill=fill("FCE4D6"),font=font(size=9,italic=True,color="C00000",bold=True),
    align=align("center",wrap=True),border=thin())
R+=2

# Cabeçalho das carteiras
sv(ws,R,1,"Métrica")
fmt(ws,R,1,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE),align=align("left"),border=thin())
for ci,nome in enumerate(nomes):
    col_t = 2 + ci*2
    col_o = 3 + ci*2
    cor = CORES.get(nome, "404040")
    sv(ws,R,col_t,nome)
    ws.merge_cells(start_row=R,start_column=col_t,end_row=R,end_column=col_o)
    ws.row_dimensions[R].height = 28
    fmt(ws,R,col_t,fill=fill(cor),font=font(bold=True,color=C_WHITE,size=10),
        align=align("center",wrap=True),border=thin())
R+=1

sv(ws,R,1,"")
fmt(ws,R,1,fill=fill(C_NAVY),border=thin())
for ci,nome in enumerate(nomes):
    col_t = 2 + ci*2; col_o = 3 + ci*2
    cor = CORES.get(nome,"404040")
    sv(ws,R,col_t,"Rec. (OOS)")
    fmt(ws,R,col_t,fill=fill(cor),font=font(bold=True,color=C_WHITE,size=9),
        align=align("center"),border=thin())
    sv(ws,R,col_o,"Otim. (OOS)")
    fmt(ws,R,col_o,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE,size=9),
        align=align("center"),border=thin())
R+=1

METRICAS_BT = [
    ("Nº de Ativos",          None,          "0",      None),
    ("─── PERÍODO COMPLETO (full-sample) ───", "__sep__", "", None),
    ("Retorno Anual (full)",   "Ra",          "0.00%",  "high"),
    ("Sharpe (full)",          "sharpe",      "0.00",   "high"),
    ("─── TESTE OUT-OF-SAMPLE ───",           "__sep__", "", None),
    ("Retorno Anual (teste)",  "Ra",          "0.00%",  "high"),
    ("Volatilidade (teste)",   "vol",         "0.00%",  "low"),
    ("Sharpe (teste)",         "sharpe",      "0.00",   "high"),
    ("Sortino (teste)",        "sortino",     "0.00",   "high"),
    ("Beta vs IBOV (teste)",   "beta",        "0.00",   None),
    ("Alpha Jensen (teste)",   "alpha_j",     "0.00%",  "high"),
    ("MDD (teste)",            "mdd",         "0.00%",  "abs_low"),
    ("VaR 95% (teste)",        "var95",       "0.000%", "abs_low"),
    ("CVaR 95% (teste)",       "cvar95",      "0.000%", "abs_low"),
    ("Retorno Acumulado (teste)", "cum_ret",  "0.00%",  "high"),
    ("─── ANÁLISE DO BACKTEST ───",           "__sep__", "", None),
    ("Sharpe Otim. IN-SAMPLE", "__train_sharpe__", "0.00", "high"),
    ("Sharpe Otim. OUT-SAMPLE","__test_sharpe__",  "0.00", "high"),
    ("Degradação Sharpe (OOS−IS)", "__deg__", "+0.00",  "abs_low"),
]

for midx, (label, key, nf, best_rule) in enumerate(METRICAS_BT):
    if key == "__sep__":
        sv(ws,R,1,label)
        ws.merge_cells(start_row=R,start_column=1,end_row=R,end_column=len(nomes)*2+2)
        ws.row_dimensions[R].height = 16
        fmt(ws,R,1,fill=fill(C_TEAL),font=font(bold=True,color=C_WHITE,size=9),
            align=align("center"),border=thin())
        R+=1; continue

    bg = "F5F5F5" if midx%2==0 else C_WHITE

    sv(ws,R,1,label)
    fmt(ws,R,1,fill=fill(bg),font=font(bold=True,size=10),align=align("left"),border=thin())

    vals_rec=[]; vals_opt=[]
    for nome in nomes:
        res = RESULTADOS[nome]
        if key is None:
            vals_rec.append(len(res["tks"])); vals_opt.append(len(res["tks"]))
        elif key == "__train_sharpe__":
            vals_rec.append(res["m_opt_train"]["sharpe"]); vals_opt.append(res["m_opt_train"]["sharpe"])
        elif key == "__test_sharpe__":
            vals_rec.append(res["m_opt_test"]["sharpe"]); vals_opt.append(res["m_opt_test"]["sharpe"])
        elif key == "__deg__":
            vals_rec.append(res["degradacao"]); vals_opt.append(res["degradacao"])
        elif "__sep__" in (key or ""):
            vals_rec.append(None); vals_opt.append(None)
        else:
            # rec usa m_rec_test / full; opt usa m_opt_test
            if "full" in label.lower():
                vals_rec.append(res["m_rec_full"].get(key, None))
                vals_opt.append(res["m_rec_full"].get(key, None))
            else:
                vals_rec.append(res["m_rec_test"].get(key, None))
                vals_opt.append(res["m_opt_test"].get(key, None))

    def bv(vals,rule):
        clean = [v for v in vals if v is not None]
        if not clean: return None
        if rule=="high":    return max(clean)
        if rule=="low":     return min(clean)
        if rule=="abs_low": return min(clean,key=abs)
        return None

    bv_rec = bv(vals_rec, best_rule)
    bv_opt = bv(vals_opt, best_rule)

    for ci,nome in enumerate(nomes):
        col_t = 2+ci*2; col_o = 3+ci*2
        vr = vals_rec[ci]; vo = vals_opt[ci]
        sv(ws,R,col_t,vr); sv(ws,R,col_o,vo)
        if vr is not None: ws.cell(R,col_t).number_format = nf
        if vo is not None: ws.cell(R,col_o).number_format = nf

        bg_r = C_GREEN if (bv_rec is not None and vr is not None and abs(vr-bv_rec)<1e-9) else bg
        bg_o = C_GREEN if (bv_opt is not None and vo is not None and abs(vo-bv_opt)<1e-9) else bg
        # Degradação: verde = pequena, vermelho = grande
        if key == "__deg__":
            bg_r = C_GREEN if vr is not None and vr > -0.3 else "FCE4D6"
            bg_o = bg_r

        fmt(ws,R,col_t,fill=fill(bg_r),font=font(size=10),align=align("center"),border=thin())
        fmt(ws,R,col_o,fill=fill(bg_o),font=font(size=10),align=align("center"),border=thin())
    R+=1

# ──────────────────────────────────────────────────────────────────
# ABA 2: BENCHMARK DIVO11
# ──────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Benchmark DIVO11")
ws2.column_dimensions["A"].width = 30
for i in range(1, len(nomes)+4):
    ws2.column_dimensions[get_column_letter(i+1)].width = 14

R2 = 1
sv(ws2,R2,1,"BENCHMARK: DIVO11 (proxy IDIV) · Carteiras de Dividendos · Junho 2026")
ws2.merge_cells(start_row=R2,start_column=1,end_row=R2,end_column=len(nomes)+3)
ws2.row_dimensions[R2].height = 22
fmt(ws2,R2,1,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE,size=12),
    align=align("center"),border=thin())
R2+=1

disc2=(f"Período completo: {df_all.index[0].date()} → {df_all.index[-1].date()} ({N_TOTAL} pregões) · "
       f"RF = {RF_AA:.2%} a.a. · DIVO11 = ETF que replica o IDIV (Índice de Dividendos B3)")
sv(ws2,R2,1,disc2)
ws2.merge_cells(start_row=R2,start_column=1,end_row=R2,end_column=len(nomes)+3)
ws2.row_dimensions[R2].height = 24
fmt(ws2,R2,1,fill=fill(C_YELLOW),font=font(size=9,italic=True,color=C_ORANGE,bold=True),
    align=align("center",wrap=True),border=thin())
R2+=1

warn2=("⚠ DIVO11 replica IDIV mas não é idêntico: há come-cotas, tracking error e taxa de adm. "
       "Alpha positivo vs DIVO11 indica geração de valor sobre o índice de dividendos, mas pode refletir "
       "seleção de ativos ex-post e não replicar fora da janela observada.")
sv(ws2,R2,1,warn2)
ws2.merge_cells(start_row=R2,start_column=1,end_row=R2,end_column=len(nomes)+3)
ws2.row_dimensions[R2].height = 28
fmt(ws2,R2,1,fill=fill("FCE4D6"),font=font(size=9,italic=True,color="C00000",bold=True),
    align=align("center",wrap=True),border=thin())
R2+=2

# Cabeçalho
todos = ["DIVO11","IBOV"] + nomes
sv(ws2,R2,1,"Métrica")
fmt(ws2,R2,1,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE),align=align("left"),border=thin())
for ci,nome in enumerate(todos):
    col = 2+ci
    cor = CORES.get(nome,"404040")
    sv(ws2,R2,col,nome)
    ws2.row_dimensions[R2].height = 24
    fmt(ws2,R2,col,fill=fill(cor),font=font(bold=True,color=C_WHITE,size=9 if len(nome)>12 else 10),
        align=align("center",wrap=True),border=thin())
R2+=1

# Construir m_todos
m_todos_full = {}
m_todos_test = {}
m_todos_full["DIVO11"] = m_divo_full
m_todos_test["DIVO11"] = m_divo_test
m_todos_full["IBOV"]   = m_ibov_full
m_todos_test["IBOV"]   = m_ibov_test
for nome,res in RESULTADOS.items():
    m_todos_full[nome] = res["m_rec_full"]
    m_todos_test[nome] = res["m_rec_test"]

METRICAS_BM = [
    ("─── PERÍODO COMPLETO (full-sample) ───", "__sep__", ""),
    ("Retorno Anual",         "Ra",      "0.00%"),
    ("Volatilidade",          "vol",     "0.00%"),
    ("Sharpe Ratio",          "sharpe",  "0.00"),
    ("Sortino Ratio",         "sortino", "0.00"),
    ("Beta vs IBOV",          "beta",    "0.00"),
    ("Alpha Jensen vs IBOV",  "alpha_j", "0.00%"),
    ("IR vs DIVO11",          "ir_divo", "0.00"),
    ("MDD",                   "mdd",     "0.00%"),
    ("Retorno Acumulado",     "cum_ret", "0.00%"),
    ("─── PERÍODO DE TESTE OOS ───",         "__sep__", ""),
    ("Retorno Anual (teste)", "Ra",      "0.00%"),
    ("Sharpe (teste)",        "sharpe",  "0.00"),
    ("Alpha Jensen (teste)",  "alpha_j", "0.00%"),
    ("IR vs DIVO11 (teste)",  "ir_divo", "0.00"),
    ("MDD (teste)",           "mdd",     "0.00%"),
]

for midx,(label,key,nf) in enumerate(METRICAS_BM):
    if key=="__sep__":
        sv(ws2,R2,1,label)
        ws2.merge_cells(start_row=R2,start_column=1,end_row=R2,end_column=len(todos)+2)
        ws2.row_dimensions[R2].height=16
        fmt(ws2,R2,1,fill=fill(C_TEAL),font=font(bold=True,color=C_WHITE,size=9),
            align=align("center"),border=thin())
        R2+=1; continue

    bg="F5F5F5" if midx%2==0 else C_WHITE
    sv(ws2,R2,1,label)
    fmt(ws2,R2,1,fill=fill(bg),font=font(bold=True,size=10),align=align("left"),border=thin())

    use_test = "teste" in label.lower()
    m_map = m_todos_test if use_test else m_todos_full

    vals = [m_map[n].get(key) for n in todos]
    # Best: Sharpe/Sortino/Ra/cum_ret/IR/alpha_j → high; vol/mdd/var95 → low/abs_low
    high_keys = {"Ra","sharpe","sortino","cum_ret","ir_divo","alpha_j"}
    rule = "high" if key in high_keys else "abs_low"
    clean = [v for v in vals if v is not None]
    bval = (max(clean) if rule=="high" else min(clean,key=abs)) if clean else None

    for ci,nome in enumerate(todos):
        col=2+ci
        v=vals[ci]
        sv(ws2,R2,col,v)
        if v is not None: ws2.cell(R2,col).number_format=nf
        is_best = bval is not None and v is not None and abs(v-bval)<1e-9
        # Destaque DIVO11 e IBOV em roxo/cinza (benchmarks não competem)
        if nome in ("DIVO11","IBOV"):
            bg_c = CORES[nome]+"22" if len(CORES.get(nome,""))<=6 else bg
            fmt(ws2,R2,col,fill=fill("E8E8E8"),font=font(size=10,italic=True),
                align=align("center"),border=thin(),nf=nf)
        else:
            bg_c = C_GREEN if is_best else bg
            fmt(ws2,R2,col,fill=fill(bg_c),font=font(size=10),
                align=align("center"),border=thin())
    R2+=1

# ──────────────────────────────────────────────────────────────────
# ABA 3: RESUMO DOS PESOS OOS
# ──────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Pesos Otimizados Treino")
ws3.column_dimensions["A"].width = 20
for ci in range(len(nomes)+3):
    ws3.column_dimensions[get_column_letter(ci+2)].width = 14

R3=1
sv(ws3,R3,1,"PESOS OTIMIZADOS (TREINO) — aplicados out-of-sample")
ws3.merge_cells(start_row=R3,start_column=1,end_row=R3,end_column=len(nomes)*3+2)
ws3.row_dimensions[R3].height=22
fmt(ws3,R3,1,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE,size=12),
    align=align("center"),border=thin())
R3+=1

disc3=("Pesos otimizados (Max Sharpe) na janela de treino e avaliados out-of-sample. "
       f"Bounds: α·wᵣₑc ≤ wᵢ ≤ β·wᵣₑc  (α={ALPHA}, β={BETA}). "
       "Verde = peso subiu vs. recomendado. Vermelho = peso caiu.")
sv(ws3,R3,1,disc3)
ws3.merge_cells(start_row=R3,start_column=1,end_row=R3,end_column=len(nomes)*3+2)
ws3.row_dimensions[R3].height=24
fmt(ws3,R3,1,fill=fill(C_YELLOW),font=font(size=9,italic=True,color=C_ORANGE,bold=True),
    align=align("center",wrap=True),border=thin())
R3+=2

for nome,res in RESULTADOS.items():
    cor=CORES.get(nome,"404040")
    sv(ws3,R3,1,nome)
    ws3.merge_cells(start_row=R3,start_column=1,end_row=R3,end_column=4)
    ws3.row_dimensions[R3].height=20
    fmt(ws3,R3,1,fill=fill(cor),font=font(bold=True,color=C_WHITE,size=11),
        align=align("center"),border=thin())
    R3+=1
    for ci,h in enumerate(["Ativo","Peso Rec.","Lb (α·rec)","Ub (β·rec)","Peso Otim. Treino","Δ"]):
        sv(ws3,R3,ci+1,h)
        fmt(ws3,R3,ci+1,fill=fill("2E75B6"),font=font(bold=True,color=C_WHITE,size=9),
            align=align("center"),border=thin())
    ws3.column_dimensions["B"].width=12; ws3.column_dimensions["C"].width=13
    ws3.column_dimensions["D"].width=13; ws3.column_dimensions["E"].width=18
    ws3.column_dimensions["F"].width=12
    R3+=1
    for i,(tk) in enumerate(res["tks"]):
        wr=float(res["w_rec"][i]); wo=float(res["w_opt_train"][i])
        d=wo-wr; lb=ALPHA*wr; ub=BETA*wr
        bg="F5F5F5" if i%2==0 else C_WHITE
        sv(ws3,R3,1,tk); sv(ws3,R3,2,wr); sv(ws3,R3,3,lb)
        sv(ws3,R3,4,ub); sv(ws3,R3,5,wo); sv(ws3,R3,6,d)
        for ci in range(1,7):
            fmt(ws3,R3,ci,fill=fill(bg),font=font(size=10,bold=(ci==1)),
                align=align("left" if ci==1 else "center"),border=thin())
            if ci>1: ws3.cell(R3,ci).number_format="0.00%"
        if   d> 0.001: fmt(ws3,R3,6,fill=fill(C_GREEN))
        elif d<-0.001: fmt(ws3,R3,6,fill=fill(C_RED))
        R3+=1
    R3+=1

wb.save(OUT_PATH)
print(f"\n✓ Arquivo salvo: {OUT_PATH}")

# ══════════════════════════════════════════════════════════════════════
# 4. RESUMO NO TERMINAL
# ══════════════════════════════════════════════════════════════════════
print(f"\n{'='*80}")
print(f"  BACKTEST OUT-OF-SAMPLE — Sharpe Otimizado (treino → teste)")
print(f"{'='*80}")
print(f"  {'Carteira':<22} {'IS (treino)':>12} {'OOS (teste)':>12} {'Degradação':>12} {'Rec. OOS':>10}")
print(f"  {'-'*70}")
for nome,res in RESULTADOS.items():
    si=res["m_opt_train"]["sharpe"]; so=res["m_opt_test"]["sharpe"]
    dr=res["degradacao"]; rs=res["m_rec_test"]["sharpe"]
    print(f"  {nome:<22} {si:>12.3f} {so:>12.3f} {dr:>12.3f} {rs:>10.3f}")

print(f"\n{'='*80}")
print(f"  BENCHMARK — Retorno Anual e Sharpe (período completo, como recomendado)")
print(f"{'='*80}")
print(f"  {'Carteira':<22} {'Retorno Anual':>14} {'Sharpe':>8} {'IR vs DIVO11':>14} {'Alpha Jensen':>13}")
print(f"  {'-'*75}")
for ref,m in [("DIVO11", m_divo_full), ("IBOV", m_ibov_full)]:
    ra=m["Ra"]; sh=m["sharpe"]
    print(f"  {ref:<22} {ra:>14.1%} {sh:>8.3f}  {'(benchmark)':>14} {'(benchmark)':>13}")
for nome,res in RESULTADOS.items():
    m=res["m_rec_full"]; ra=m["Ra"]; sh=m["sharpe"]
    ir=m.get("ir_divo") or 0.0; al=m.get("alpha_j") or 0.0
    print(f"  {nome:<22} {ra:>14.1%} {sh:>8.3f} {ir:>14.3f} {al:>13.1%}")

print(f"\n⚠ Análise retroativa. Não é recomendação de investimento.")
print(f"  Desempenho passado não garante desempenho futuro.")

# -*- coding: utf-8 -*-
"""
monte_carlo.py
══════════════════════════════════════════════════════════════════════
Módulo A — BOOTSTRAP BACKTEST (500 splits aleatórios)
  - Sorteia ponto de corte aleatório entre índices 80 e 168
  - Otimiza Max Sharpe no treino; avalia no teste
  - Distribui Sharpe OOS: P5, mediana, P95, % splits com Sharpe > 0
  - Responde: "o otimizador adiciona valor de forma consistente?"

Módulo B — MONTE CARLO FORWARD (5000 trajetórias, 252 pregões)
  - Parâmetros: mu e cov históricos do período completo
  - Simulação por Cholesky (correlações preservadas)
  - P5 / P50 / P95 do retorno acumulado em 1 ano
  - Probabilidade de superar RF, DIVO11 e de MDD > 20%

══════════════════════════════════════════════════════════════════════
DISCLAIMERS:
  • Análise retroativa e prospectiva hipotética.
  • NÃO é recomendação de investimento.
  • Bootstrap usa janelas contíguas — não captura breaks estruturais.
  • Monte Carlo assume retornos normais multivariados — caudas reais
    são mais pesadas (fat tails). Resultados são indicativos.
  • Desempenho passado não garante desempenho futuro.
══════════════════════════════════════════════════════════════════════
"""

import os, math, warnings
import numpy as np, pandas as pd, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from scipy.optimize import minimize, Bounds

warnings.filterwarnings("ignore")
np.random.seed(42)

# ══════════════════════════════════════════════════════════════════════
# PARÂMETROS
# ══════════════════════════════════════════════════════════════════════
PASTA    = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(PASTA, "precos_b3.csv")
OUT_PATH = os.path.join(PASTA, "monte_carlo.xlsx")

RF_AA     = 0.1440
RF_DIARIO = math.log(1 + RF_AA) / 252
ALPHA = 0.5; BETA = 2.0
N_BOOTSTRAP = 500
N_MC        = 5000
H_MC        = 252   # horizonte forward (1 ano útil)
MIN_TREINO  = 80    # mínimo de pregões em cada janela

CARTEIRAS = {
    "XP Dividendos": {
        "PETR4":0.100,"PRIO3":0.050,"VALE3":0.125,"CPLE3":0.150,
        "ENGI11":0.050,"AXIA3":0.100,"VIVT3":0.050,"ALOS3":0.075,
        "B3SA3":0.100,"CXSE3":0.050,"ITUB4":0.150,
    },
    "BTG Dividendos": {
        "PETR4":0.10,"ITUB4":0.10,"VALE3":0.05,"BBDC4":0.10,
        "AXIA3":0.10,"EQTL3":0.10,"CPLE3":0.10,"CXSE3":0.10,
        "MOTV3":0.05,"CSMG3":0.05,"ALOS3":0.05,"CURY3":0.10,
    },
    "Itau BBA Div.": {
        "AXIA3":0.20,"ALOS3":0.20,"BBDC4":0.20,"VALE3":0.20,"PETR4":0.20,
    },
    "Santander Div.": {
        "ALOS3":0.10,"AXIA3":0.10,"BPAC11":0.10,"CPLE3":0.10,
        "VIVT3":0.10,"VALE3":0.10,"VBBR3":0.10,"CURY3":0.10,
        "ITUB4":0.10,"PETR3":0.10,
    },
    "BB Dividendos": {
        "ALOS3":0.10,"ABEV3":0.10,"BBDC4":0.10,"BRAP4":0.10,
        "CXSE3":0.10,"DIRR3":0.10,"ITSA4":0.10,"PETR4":0.10,
        "TAEE11":0.10,"TIMS3":0.10,
    },
}
CORES = {
    "XP Dividendos":"C55A11","BTG Dividendos":"17375E","Itau BBA Div.":"1E5799",
    "Santander Div.":"C00000","BB Dividendos":"375623",
    "DIVO11":"7030A0","IBOV":"404040",
}
C_NAVY="1F3864";C_WHITE="FFFFFF";C_YELLOW="FFF2CC";C_GREEN="C6EFCE"
C_RED="FCE4D6";C_TEAL="1E6B6B";C_ORANGE="C55A11";C_GRAY="F5F5F5"

def fill(h): return PatternFill("solid",fgColor=h)
def font(bold=False,color="1A1A2E",size=10,italic=False):
    return Font(bold=bold,color=color,size=size,italic=italic,name="Calibri")
def align(h="center",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def thin(c="BFBFBF"):
    s=Side(style="thin",color=c); return Border(left=s,right=s,top=s,bottom=s)
def sv(ws,r,c,v):
    cell=ws.cell(r,c)
    if not isinstance(cell,MergedCell): cell.value=v
def fmt(ws,r,c,**kw):
    cell=ws.cell(r,c)
    if isinstance(cell,MergedCell): return
    for k,v in kw.items():
        if   k=="fill":   cell.fill=v
        elif k=="font":   cell.font=v
        elif k=="align":  cell.alignment=v
        elif k=="border": cell.border=v
        elif k=="nf":     cell.number_format=v

# ══════════════════════════════════════════════════════════════════════
# FUNÇÕES CORE
# ══════════════════════════════════════════════════════════════════════
def log_rets_arr(df_sub, tks):
    df = df_sub[tks].ffill(limit=5)
    arr = df.values.astype(float)
    lr  = np.where(np.isnan(arr[1:])|np.isnan(arr[:-1]), 0.0,
                   np.log(np.clip(arr[1:]/arr[:-1], 1e-8, None)))
    return np.nan_to_num(lr)

def log_rets_serie(df_sub, col):
    s = df_sub[col].ffill(limit=5).values.astype(float)
    return np.where(np.isnan(s[1:])|np.isnan(s[:-1]), 0.0,
                    np.log(np.clip(s[1:]/s[:-1], 1e-8, None)))

def sharpe_portfolio(w, lr, ibov_lr=None):
    pr  = lr @ w
    Ra  = float(np.mean(pr) * 252)
    vol = float(np.std(pr, ddof=1) * math.sqrt(252))
    return (Ra - RF_AA) / vol if vol > 1e-12 else -999.0

def otimizar(w_rec, lr):
    lr = np.nan_to_num(lr)
    EPS = 1e-4
    lb  = np.maximum(ALPHA * w_rec, EPS)
    ub  = np.minimum(BETA  * w_rec, 1.0)
    if np.sum(lb) > 1.0 + 1e-6: lb = lb / np.sum(lb) * 0.99
    ub  = np.maximum(ub, lb + 1e-5)

    mu_a  = np.mean(lr, 0) * 252
    cov_a = np.atleast_2d(np.cov(lr.T, ddof=1) * 252)

    def neg_sh(w):
        r = float(w @ mu_a)
        v = float(np.sqrt(np.clip(w @ cov_a @ w, 0, None)))
        return -(r - RF_AA) / v if v > 1e-12 else 999.0

    w0 = np.clip(w_rec.copy(), lb, ub); w0 /= w0.sum()
    res = minimize(neg_sh, w0, method="SLSQP",
                   bounds=Bounds(lb=lb, ub=ub),
                   constraints=[{"type":"eq","fun":lambda w: w.sum()-1.0}],
                   options={"maxiter":3000,"ftol":1e-12})
    w_opt = np.clip(res.x, lb, ub); w_opt /= w_opt.sum()
    return w_opt

# ══════════════════════════════════════════════════════════════════════
# 1. CARREGAR DADOS
# ══════════════════════════════════════════════════════════════════════
print("═"*65)
print("MONTE CARLO — Carteiras de Dividendos · Junho 2026")
print("═"*65)

df_all = pd.read_csv(CSV_PATH, index_col=0, parse_dates=True).sort_index().dropna(how="all")
N = len(df_all)
print(f"\nDados: {df_all.index[0].date()} → {df_all.index[-1].date()} ({N} pregões)")
print(f"Bootstrap: {N_BOOTSTRAP} splits aleatórios | Monte Carlo: {N_MC} trajetórias × {H_MC} pregões\n")

# ══════════════════════════════════════════════════════════════════════
# MÓDULO A — BOOTSTRAP BACKTEST
# ══════════════════════════════════════════════════════════════════════
print("Módulo A — Bootstrap Backtest...")

split_min = MIN_TREINO
split_max = N - MIN_TREINO - 1
splits    = np.random.randint(split_min, split_max+1, size=N_BOOTSTRAP)

BOOT = {}
for nome, pesos in CARTEIRAS.items():
    tks    = list(pesos.keys())
    w_rec  = np.array(list(pesos.values()), dtype=float); w_rec /= w_rec.sum()
    sh_rec_oos  = []
    sh_opt_oos  = []
    sh_opt_is   = []

    for idx, sp in enumerate(splits):
        df_tr = df_all.iloc[:sp+1]
        df_te = df_all.iloc[sp+1:]
        if len(df_te) < MIN_TREINO: continue

        lr_tr = log_rets_arr(df_tr, tks)
        lr_te = log_rets_arr(df_te, tks)
        if lr_tr.shape[0] < 30 or lr_te.shape[0] < 20: continue

        w_opt = otimizar(w_rec, lr_tr)

        sh_rec_oos.append(sharpe_portfolio(w_rec, lr_te))
        sh_opt_is .append(sharpe_portfolio(w_opt, lr_tr))
        sh_opt_oos.append(sharpe_portfolio(w_opt, lr_te))

    sh_rec_oos  = np.array(sh_rec_oos)
    sh_opt_oos  = np.array(sh_opt_oos)
    sh_opt_is   = np.array(sh_opt_is)
    delta       = sh_opt_oos - sh_rec_oos

    BOOT[nome] = {
        "sh_rec_oos": sh_rec_oos, "sh_opt_oos": sh_opt_oos,
        "sh_opt_is":  sh_opt_is,  "delta":       delta,
        "n": len(sh_rec_oos),
        # resumo recomendado
        "rec_p5":  np.percentile(sh_rec_oos, 5),
        "rec_p50": np.median(sh_rec_oos),
        "rec_p95": np.percentile(sh_rec_oos, 95),
        "rec_pct_pos": np.mean(sh_rec_oos > 0),
        # resumo otimizado IS
        "is_p50": np.median(sh_opt_is),
        # resumo otimizado OOS
        "opt_p5":  np.percentile(sh_opt_oos, 5),
        "opt_p50": np.median(sh_opt_oos),
        "opt_p95": np.percentile(sh_opt_oos, 95),
        "opt_pct_pos": np.mean(sh_opt_oos > 0),
        # valor do otimizador
        "delta_p50": np.median(delta),
        "delta_pct_pos": np.mean(delta > 0),
    }
    print(f"  {nome:<22}  Rec OOS mediana={BOOT[nome]['rec_p50']:+.3f}  "
          f"Opt OOS mediana={BOOT[nome]['opt_p50']:+.3f}  "
          f"Δ mediana={BOOT[nome]['delta_p50']:+.3f}  "
          f"Opt>Rec: {BOOT[nome]['delta_pct_pos']:.0%}")

# ══════════════════════════════════════════════════════════════════════
# MÓDULO B — MONTE CARLO FORWARD
# ══════════════════════════════════════════════════════════════════════
print("\nMódulo B — Monte Carlo Forward...")

# Benchmark DIVO11 para comparação
divo_lr_full = log_rets_serie(df_all, "DIVO11")
divo_mu = float(np.mean(divo_lr_full) * 252)   # retorno anual DIVO11
divo_vol= float(np.std(divo_lr_full, ddof=1) * math.sqrt(252))

MC = {}
for nome, pesos in CARTEIRAS.items():
    tks   = list(pesos.keys())
    w_rec = np.array(list(pesos.values()), dtype=float); w_rec /= w_rec.sum()

    # Parâmetros históricos completos
    lr_full = log_rets_arr(df_all, tks)
    mu_d    = np.mean(lr_full, 0)    # média diária por ativo
    cov_d   = np.atleast_2d(np.cov(lr_full.T, ddof=1))

    # Decomposição de Cholesky para simular retornos correlacionados
    try:
        L = np.linalg.cholesky(cov_d + np.eye(len(tks))*1e-8)
    except np.linalg.LinAlgError:
        L = np.linalg.cholesky(cov_d + np.eye(len(tks))*1e-6)

    # mu da carteira (diário) e vol
    mu_port  = float(mu_d @ w_rec)
    vol_port = float(np.sqrt(w_rec @ cov_d @ w_rec))

    # Simular N_MC caminhos
    # shape: (N_MC, H_MC, n_ativos)
    Z    = np.random.standard_normal((N_MC, H_MC, len(tks)))
    shocks = Z @ L.T + mu_d   # (N_MC, H_MC, n_ativos)
    ret_port = shocks @ w_rec  # (N_MC, H_MC)

    # Riqueza acumulada (base 1)
    wealth = np.exp(np.cumsum(ret_port, axis=1))  # (N_MC, H_MC)
    final_w= wealth[:, -1]                         # terminal

    # Retorno acumulado anual
    ret_acum = final_w - 1.0

    # Sharpe simulado de cada trajetória
    Ra_traj  = np.mean(ret_port, axis=1) * 252
    vol_traj = np.std(ret_port, axis=1, ddof=1) * math.sqrt(252)
    sh_traj  = np.where(vol_traj > 1e-12, (Ra_traj - RF_AA) / vol_traj, -999.0)

    # MDD de cada trajetória
    peak_w   = np.maximum.accumulate(wealth, axis=1)
    dd       = (wealth - peak_w) / peak_w
    mdd_traj = np.min(dd, axis=1)

    # Probabilidades
    p_gt_rf    = np.mean(ret_acum > RF_AA)          # supera RF a.a.
    p_gt_divo  = np.mean(Ra_traj > divo_mu)         # supera DIVO11 anual
    p_mdd20    = np.mean(mdd_traj < -0.20)           # MDD pior que -20%

    MC[nome] = {
        "mu_port": mu_port*252, "vol_port": vol_port*math.sqrt(252),
        "ret_p5":   np.percentile(ret_acum, 5),
        "ret_p25":  np.percentile(ret_acum, 25),
        "ret_p50":  np.percentile(ret_acum, 50),
        "ret_p75":  np.percentile(ret_acum, 75),
        "ret_p95":  np.percentile(ret_acum, 95),
        "sh_p5":    np.percentile(sh_traj, 5),
        "sh_p50":   np.percentile(sh_traj, 50),
        "sh_p95":   np.percentile(sh_traj, 95),
        "mdd_p50":  np.percentile(mdd_traj, 50),
        "mdd_p5":   np.percentile(mdd_traj, 5),
        "p_gt_rf":   p_gt_rf,
        "p_gt_divo": p_gt_divo,
        "p_mdd20":   p_mdd20,
        "wealth": wealth,
    }
    print(f"  {nome:<22}  Ret mediana={MC[nome]['ret_p50']:.1%}  "
          f"[P5={MC[nome]['ret_p5']:.1%}, P95={MC[nome]['ret_p95']:.1%}]  "
          f"P(>RF)={p_gt_rf:.0%}  P(>DIVO11)={p_gt_divo:.0%}  P(MDD>20%)={p_mdd20:.0%}")

# ══════════════════════════════════════════════════════════════════════
# 2. EXCEL
# ══════════════════════════════════════════════════════════════════════
print(f"\nGerando Excel: {OUT_PATH}")
wb = openpyxl.Workbook()
nomes = list(CARTEIRAS.keys())

# ─────────────────────────────────────────────────────────────────────
# ABA 1: BOOTSTRAP BACKTEST
# ─────────────────────────────────────────────────────────────────────
ws1 = wb.active; ws1.title = "Bootstrap Backtest"
ws1.column_dimensions["A"].width = 32
for ci in range(len(nomes)+2):
    ws1.column_dimensions[get_column_letter(ci+2)].width = 14

R=1
sv(ws1,R,1,f"BOOTSTRAP BACKTEST · {N_BOOTSTRAP} splits aleatórios · Carteiras de Dividendos · Junho 2026")
ws1.merge_cells(start_row=R,start_column=1,end_row=R,end_column=len(nomes)+2)
ws1.row_dimensions[R].height=22
fmt(ws1,R,1,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE,size=12),align=align("center"),border=thin())
R+=1

disc = (f"Cada split: ponto de corte aleatório ∈ [{MIN_TREINO}..{N-MIN_TREINO-1}] pregões. "
        f"Otimização Max Sharpe com α={ALPHA}, β={BETA}. "
        "Sharpe Recomendado OOS = pesos originais das casas avaliados fora da amostra. "
        "Sharpe Otimizado OOS = pesos otimizados no treino avaliados no teste.")
sv(ws1,R,1,disc)
ws1.merge_cells(start_row=R,start_column=1,end_row=R,end_column=len(nomes)+2)
ws1.row_dimensions[R].height=28
fmt(ws1,R,1,fill=fill(C_YELLOW),font=font(size=9,italic=True,color=C_ORANGE,bold=True),
    align=align("center",wrap=True),border=thin())
R+=2

# Cabeçalho
sv(ws1,R,1,"Métrica"); fmt(ws1,R,1,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE),align=align("left"),border=thin())
for ci,nome in enumerate(nomes):
    cor=CORES.get(nome,"404040"); col=2+ci
    sv(ws1,R,col,nome); ws1.row_dimensions[R].height=24
    fmt(ws1,R,col,fill=fill(cor),font=font(bold=True,color=C_WHITE,size=9),align=align("center",wrap=True),border=thin())
R+=1

secs = [
    ("─── PESOS RECOMENDADOS (OOS) ───", "__sep__",""),
    ("Sharpe OOS — P5  (pior 5%)",      "rec_p5",  "+0.00"),
    ("Sharpe OOS — Mediana",            "rec_p50", "+0.00"),
    ("Sharpe OOS — P95 (melhor 5%)",    "rec_p95", "+0.00"),
    ("% splits com Sharpe > 0",         "rec_pct_pos","0%"),
    ("─── PESOS OTIMIZADOS (OOS) ───",  "__sep__",""),
    ("Sharpe IS (treino) — Mediana",    "is_p50",  "+0.00"),
    ("Sharpe OOS — P5  (pior 5%)",      "opt_p5",  "+0.00"),
    ("Sharpe OOS — Mediana",            "opt_p50", "+0.00"),
    ("Sharpe OOS — P95 (melhor 5%)",    "opt_p95", "+0.00"),
    ("% splits com Sharpe > 0",         "opt_pct_pos","0%"),
    ("─── VALOR DO OTIMIZADOR ───",     "__sep__",""),
    ("Δ Sharpe (Opt − Rec) Mediana",    "delta_p50","+0.00"),
    ("% splits Opt > Rec",              "delta_pct_pos","0%"),
]

for midx,(label,key,nf) in enumerate(secs):
    if key=="__sep__":
        sv(ws1,R,1,label)
        ws1.merge_cells(start_row=R,start_column=1,end_row=R,end_column=len(nomes)+2)
        ws1.row_dimensions[R].height=16
        fmt(ws1,R,1,fill=fill(C_TEAL),font=font(bold=True,color=C_WHITE,size=9),align=align("center"),border=thin())
        R+=1; continue
    bg="F5F5F5" if midx%2==0 else C_WHITE
    sv(ws1,R,1,label); fmt(ws1,R,1,fill=fill(bg),font=font(bold=True,size=10),align=align("left"),border=thin())
    vals=[BOOT[n][key] for n in nomes]
    # best: mediana e pct_pos → max; p5 é informativo; delta_p50 → se positivo é bom
    high_keys={"rec_p50","rec_p95","rec_pct_pos","is_p50","opt_p50","opt_p95","opt_pct_pos","delta_p50","delta_pct_pos"}
    if key in high_keys:
        bval=max(vals)
    else:
        bval=None
    for ci,nome in enumerate(nomes):
        col=2+ci; v=vals[ci]
        sv(ws1,R,col,v); ws1.cell(R,col).number_format=nf
        is_best=(bval is not None and abs(v-bval)<1e-9)
        if key=="delta_p50":
            bg_c=C_GREEN if v>0.05 else (C_RED if v<-0.05 else bg)
        elif key=="delta_pct_pos":
            bg_c=C_GREEN if v>0.55 else (C_RED if v<0.45 else bg)
        else:
            bg_c=C_GREEN if is_best else bg
        fmt(ws1,R,col,fill=fill(bg_c),font=font(size=10),align=align("center"),border=thin())
    R+=1

# Veredicto automático
R+=1
sv(ws1,R,1,"INTERPRETAÇÃO AUTOMÁTICA")
ws1.merge_cells(start_row=R,start_column=1,end_row=R,end_column=len(nomes)+2)
ws1.row_dimensions[R].height=18
fmt(ws1,R,1,fill=fill("2E75B6"),font=font(bold=True,color=C_WHITE,size=10),align=align("center"),border=thin())
R+=1

for nome in nomes:
    b=BOOT[nome]
    d50=b["delta_p50"]; dpct=b["delta_pct_pos"]; opt50=b["opt_p50"]; rec50=b["rec_p50"]
    if dpct>0.55 and d50>0.05:
        verd=f"Otimizador AGREGA valor ({dpct:.0%} dos splits melhora; Δ mediana={d50:+.3f})"
        cor_v="375623"
    elif dpct<0.45 or d50<-0.05:
        verd=f"Otimizador NAO agrega valor sistematicamente ({dpct:.0%} dos splits melhora; Δ mediana={d50:+.3f})"
        cor_v="C00000"
    else:
        verd=f"Resultado INCONCLUSIVO — otimizador melhora em {dpct:.0%} dos splits (Δ mediana={d50:+.3f})"
        cor_v="7F6000"
    sv(ws1,R,1,f"{nome}: {verd}")
    ws1.merge_cells(start_row=R,start_column=1,end_row=R,end_column=len(nomes)+2)
    ws1.row_dimensions[R].height=18
    fmt(ws1,R,1,fill=fill("F2F2F2"),font=font(size=10,color=cor_v,bold=True),align=align("left"),border=thin())
    R+=1

# ─────────────────────────────────────────────────────────────────────
# ABA 2: MONTE CARLO FORWARD
# ─────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Monte Carlo Forward")
ws2.column_dimensions["A"].width = 36
for ci in range(len(nomes)+2):
    ws2.column_dimensions[get_column_letter(ci+2)].width = 14

R2=1
sv(ws2,R2,1,f"MONTE CARLO FORWARD · {N_MC:,} trajetórias × {H_MC} pregões (≈ 1 ano) · Junho 2026")
ws2.merge_cells(start_row=R2,start_column=1,end_row=R2,end_column=len(nomes)+2)
ws2.row_dimensions[R2].height=22
fmt(ws2,R2,1,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE,size=12),align=align("center"),border=thin())
R2+=1

disc2=(f"Parâmetros: μ e Σ históricos (Jun/2025–Mai/2026). Simulação Cholesky (retornos normais multivariados). "
       f"Pesos recomendados pelas casas. RF = {RF_AA:.2%} a.a. | DIVO11 retorno histórico anual = {divo_mu:.1%}. "
       "Fat tails reais excedem o normal — P5 pode ser otimista.")
sv(ws2,R2,1,disc2)
ws2.merge_cells(start_row=R2,start_column=1,end_row=R2,end_column=len(nomes)+2)
ws2.row_dimensions[R2].height=28
fmt(ws2,R2,1,fill=fill(C_YELLOW),font=font(size=9,italic=True,color=C_ORANGE,bold=True),
    align=align("center",wrap=True),border=thin())
R2+=2

sv(ws2,R2,1,"Métrica"); fmt(ws2,R2,1,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE),align=align("left"),border=thin())
for ci,nome in enumerate(nomes):
    cor=CORES.get(nome,"404040"); col=2+ci
    sv(ws2,R2,col,nome); ws2.row_dimensions[R2].height=24
    fmt(ws2,R2,col,fill=fill(cor),font=font(bold=True,color=C_WHITE,size=9),align=align("center",wrap=True),border=thin())
R2+=1

secs2=[
    ("─── RETORNO ACUMULADO EM 1 ANO ───","__sep__",""),
    ("Parâmetro Histórico (μ anual)",     "mu_port",  "0.00%"),
    ("Parâmetro Histórico (σ anual)",     "vol_port", "0.00%"),
    ("Retorno P5  (cenário adverso)",     "ret_p5",   "0.00%"),
    ("Retorno P25",                        "ret_p25",  "0.00%"),
    ("Retorno P50 (cenário mediano)",     "ret_p50",  "0.00%"),
    ("Retorno P75",                        "ret_p75",  "0.00%"),
    ("Retorno P95 (cenário favorável)",   "ret_p95",  "0.00%"),
    ("─── SHARPE SIMULADO ───",          "__sep__",""),
    ("Sharpe P5",                         "sh_p5",    "+0.00"),
    ("Sharpe P50 (mediano)",              "sh_p50",   "+0.00"),
    ("Sharpe P95",                        "sh_p95",   "+0.00"),
    ("─── DRAWDOWN MÁXIMO ───",          "__sep__",""),
    ("MDD Mediano (P50)",                 "mdd_p50",  "0.00%"),
    ("MDD P5 (pior 5%)",                  "mdd_p5",   "0.00%"),
    ("─── PROBABILIDADES ───",           "__sep__",""),
    (f"P(Ret > RF = {RF_AA:.1%})",        "p_gt_rf",  "0%"),
    (f"P(Ret > DIVO11 = {divo_mu:.1%})", "p_gt_divo","0%"),
    ("P(MDD > 20%)",                      "p_mdd20",  "0%"),
]

for midx,(label,key,nf) in enumerate(secs2):
    if key=="__sep__":
        sv(ws2,R2,1,label)
        ws2.merge_cells(start_row=R2,start_column=1,end_row=R2,end_column=len(nomes)+2)
        ws2.row_dimensions[R2].height=16
        fmt(ws2,R2,1,fill=fill(C_TEAL),font=font(bold=True,color=C_WHITE,size=9),align=align("center"),border=thin())
        R2+=1; continue
    bg="F5F5F5" if midx%2==0 else C_WHITE
    sv(ws2,R2,1,label); fmt(ws2,R2,1,fill=fill(bg),font=font(bold=True,size=10),align=align("left"),border=thin())
    vals=[MC[n][key] for n in nomes]
    high_k={"mu_port","ret_p50","ret_p95","sh_p50","sh_p95","p_gt_rf","p_gt_divo"}
    low_k ={"vol_port","mdd_p50","mdd_p5","p_mdd20"}
    abs_k ={}
    if key in high_k: bval=max(vals)
    elif key in low_k: bval=min(vals,key=abs)
    else: bval=None
    for ci,nome in enumerate(nomes):
        col=2+ci; v=vals[ci]
        sv(ws2,R2,col,v); ws2.cell(R2,col).number_format=nf
        is_best=(bval is not None and abs(v-bval)<1e-9)
        if key=="p_mdd20": bg_c=C_GREEN if v<0.15 else (C_RED if v>0.25 else bg)
        else:               bg_c=C_GREEN if is_best else bg
        fmt(ws2,R2,col,fill=fill(bg_c),font=font(size=10),align=align("center"),border=thin())
    R2+=1

# ─────────────────────────────────────────────────────────────────────
# ABA 3: DISTRIBUIÇÃO DO SHARPE OOS (histograma por decis)
# ─────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Distribuicao Sharpe OOS")
ws3.column_dimensions["A"].width = 20
for ci in range(len(nomes)+2):
    ws3.column_dimensions[get_column_letter(ci+2)].width = 12

R3=1
sv(ws3,R3,1,f"DISTRIBUIÇÃO DO SHARPE OOS — {N_BOOTSTRAP} Bootstrap Splits")
ws3.merge_cells(start_row=R3,start_column=1,end_row=R3,end_column=len(nomes)+2)
ws3.row_dimensions[R3].height=20
fmt(ws3,R3,1,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE,size=12),align=align("center"),border=thin())
R3+=2

# Percentis granulares do Sharpe OOS otimizado
percs=[1,5,10,25,50,75,90,95,99]
sv(ws3,R3,1,"Percentil"); fmt(ws3,R3,1,fill=fill(C_NAVY),font=font(bold=True,color=C_WHITE),align=align("left"),border=thin())
for ci,nome in enumerate(nomes):
    sv(ws3,R3,ci+2,f"{nome[:14]}…" if len(nome)>14 else nome)
    fmt(ws3,R3,ci+2,fill=fill(CORES.get(nome,"404040")),font=font(bold=True,color=C_WHITE,size=9),
        align=align("center",wrap=True),border=thin())
R3+=1
for p in percs:
    bg="F5F5F5" if p%20==0 else C_WHITE
    lbl=f"P{p:2d}"
    sv(ws3,R3,1,lbl); fmt(ws3,R3,1,fill=fill(bg),font=font(bold=True),align=align("center"),border=thin())
    for ci,nome in enumerate(nomes):
        v=float(np.percentile(BOOT[nome]["sh_opt_oos"],p))
        sv(ws3,R3,ci+2,v); ws3.cell(R3,ci+2).number_format="+0.00"
        bg_c=C_GREEN if v>0.3 else (C_RED if v<0 else bg)
        fmt(ws3,R3,ci+2,fill=fill(bg_c),font=font(size=10),align=align("center"),border=thin())
    R3+=1

wb.save(OUT_PATH)
print(f"✓ Salvo: {OUT_PATH}")

# ══════════════════════════════════════════════════════════════════════
# 3. RESUMO TERMINAL
# ══════════════════════════════════════════════════════════════════════
print(f"\n{'═'*80}")
print(f"  BOOTSTRAP — Mediana Sharpe OOS ({N_BOOTSTRAP} splits)")
print(f"{'═'*80}")
print(f"  {'Carteira':<22} {'Rec OOS':>9} {'Opt IS':>9} {'Opt OOS':>9} {'Δ mediana':>10} {'Opt>Rec':>8} {'Veredicto'}")
print(f"  {'─'*78}")
for nome in nomes:
    b=BOOT[nome]
    d50=b["delta_p50"]; dpct=b["delta_pct_pos"]
    v="AGREGA" if dpct>0.55 and d50>0.05 else ("NAO AGREGA" if dpct<0.45 or d50<-0.05 else "INCONCLUSIVO")
    print(f"  {nome:<22} {b['rec_p50']:>9.3f} {b['is_p50']:>9.3f} {b['opt_p50']:>9.3f} "
          f"{d50:>10.3f} {dpct:>8.1%}  {v}")

print(f"\n{'═'*80}")
print(f"  MONTE CARLO — Retorno Mediano e P(>RF) em 1 ano ({N_MC:,} trajetórias)")
print(f"{'═'*80}")
print(f"  {'Carteira':<22} {'P5':>7} {'Mediana':>9} {'P95':>7} {'P(>RF)':>8} {'P(>DIVO)':>9} {'P(MDD>20%)':>11}")
print(f"  {'─'*78}")
for nome in nomes:
    m=MC[nome]
    print(f"  {nome:<22} {m['ret_p5']:>7.1%} {m['ret_p50']:>9.1%} {m['ret_p95']:>7.1%} "
          f"{m['p_gt_rf']:>8.0%} {m['p_gt_divo']:>9.0%} {m['p_mdd20']:>11.0%}")

print(f"\n  DIVO11 histórico: retorno anual={divo_mu:.1%}  vol={divo_vol:.1%}")
print(f"\n⚠ Retroativo + hipotético. Não é recomendação de investimento.")

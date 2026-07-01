# -*- coding: utf-8 -*-
"""
analisar_carteiras.py
══════════════════════════════════════════════════════════════════════
Análise comparativa e otimização de Markowitz para as carteiras
DIVIDENDOS de junho 2026 das principais casas brasileiras.

Carteiras: Casa A — Portfólio Concentrado | Casa B | Itaú BBA Dividendos
           Santander Dividendos   | BB Investimentos Dividendos

Otimização: Max Sharpe com restrições por ativo
    Σ wᵢ = 1
    α · wᵢ_rec ≤ wᵢ ≤ β · wᵢ_rec   (α = 0.5, β = 2.0)
    wᵢ > 0 (nenhum ativo zerado)

═══════════════════════════════════════════════════════════════════════
DISCLAIMERS:
  • Análise retroativa (backward-looking) – NÃO é recomendação de invest.
  • Otimização Markowitz sofre de maximização de erro de estimação.
  • Correlações sobem em crises, reduzindo a diversificação quando mais importa.
  • Desempenho passado não é garantia de desempenho futuro.
  • Dados: snapshot estático de preços ajustados – não executa chamadas de API.
═══════════════════════════════════════════════════════════════════════

Execute com a planilha FECHADA:
    python analisar_carteiras.py

Dependências: pip install openpyxl pandas numpy scipy
"""

import os, math, json
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from scipy.optimize import minimize, Bounds

# ═══════════════════════════════════════════════════════════════════
# PARÂMETROS
# ═══════════════════════════════════════════════════════════════════
PASTA    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(PASTA, "dados", "precos_b3.csv")
OUT_PATH = os.path.join(PASTA, "planilhas", "comparacao_carteiras.xlsx")

RF_AA     = 0.1440          # taxa livre de risco anual (CDI ~14.40% a.a.)
RF_DIARIO = math.log(1 + RF_AA) / 252

ALPHA = 0.5   # limite inferior: wᵢ ≥ α × wᵢ_rec
BETA  = 2.0   # limite superior: wᵢ ≤ β × wᵢ_rec

# ─── Definição das carteiras ──────────────────────────────────────
# Tickers ausentes no CSV são tratados abaixo (substituição ou exclusão)
# AXIA6 → substituído por AXIA3 (mesma empresa, classe PN vs ON)
# ROXO34 → excluído (BDR Nubank sem histórico no CSV); pesos renormalizados

CARTEIRAS = {
    "Casa A": {
        # Casa A — Portfólio Concentrado — junho 2026
        # Fonte: Money Times 02/06/2026
        # Nota: tabela original tinha AXIA7+AXIA3 (5%+5%) → AXIA7 não existe na B3; consolidado como AXIA3 10%
        "PETR4":  0.100,
        "PRIO3":  0.050,
        "VALE3":  0.125,
        "CPLE3":  0.150,
        "ENGI11": 0.050,
        "AXIA3":  0.100,
        "VIVT3":  0.050,
        "ALOS3":  0.075,
        "B3SA3":  0.100,
        "CXSE3":  0.050,
        "ITUB4":  0.150,
    },
    "Casa B": {
        # Casa B — Portfólio Equiponderado — junho 2026
        # Fonte: Money Times 01/06/2026 (tabela com pesos individuais)
        "PETR4": 0.10,
        "ITUB4": 0.10,
        "VALE3": 0.05,
        "BBDC4": 0.10,
        "AXIA3": 0.10,
        "EQTL3": 0.10,
        "CPLE3": 0.10,
        "CXSE3": 0.10,
        "MOTV3": 0.05,
        "CSMG3": 0.05,
        "ALOS3": 0.05,
        "CURY3": 0.10,
    },
    "Itaú BBA Div.": {
        # Itaú BBA Carteira Dividendos — junho 2026
        "AXIA3": 0.20,
        "ALOS3": 0.20,
        "BBDC4": 0.20,
        "VALE3": 0.20,
        "PETR4": 0.20,
    },
    "Santander Div.": {
        # Santander Carteira Dividendos — junho 2026
        # AXIA6 → substituído por AXIA3 (mesma empresa, PN vs ON)
        "ALOS3":  0.10,
        "AXIA6":  0.10,   # → substituído por AXIA3
        "BPAC11": 0.10,
        "CPLE3":  0.10,
        "VIVT3":  0.10,
        "VALE3":  0.10,
        "VBBR3":  0.10,
        "CURY3":  0.10,
        "ITUB4":  0.10,
        "PETR3":  0.10,
    },
    "BB Dividendos": {
        # BB Investimentos Carteira Dividendos — junho 2026
        # Fonte: Money Times 02/06/2026 (7 trocas vs maio)
        "ALOS3":  0.10,
        "ABEV3":  0.10,
        "BBDC4":  0.10,
        "BRAP4":  0.10,
        "CXSE3":  0.10,
        "DIRR3":  0.10,
        "ITSA4":  0.10,
        "PETR4":  0.10,
        "TAEE11": 0.10,
        "TIMS3":  0.10,
    },
}

# ─── Substituições de ticker ──────────────────────────────────────
SUBSTITUICOES = {
    "AXIA6": "AXIA3",   # PN → ON mesma empresa (Axia Energia, ex-Eletrobras)
}

# ═══════════════════════════════════════════════════════════════════
# PALETA DE CORES
# ═══════════════════════════════════════════════════════════════════
C_NAVY    = "1F3864"
C_BLUE    = "2E75B6"
C_TEAL    = "1E6B6B"
C_TEAL_LT = "D0EEEE"
C_BLUE_LT = "D6E4F0"
C_GREEN   = "C6EFCE"
C_YELLOW  = "FFF2CC"
C_RED_LT  = "FCE4D6"
C_WHITE   = "FFFFFF"
C_GRAY_LT = "F5F5F5"
C_GRAY    = "D9D9D9"
C_PURPLE  = "7030A0"
C_PURPLE_LT = "EAD1DC"
C_ORANGE  = "C55A11"

CORES_CARTEIRAS = {
    "Casa A":   "C55A11",
    "Casa B":  "17375E",
    "Itaú BBA Div.":   "1E5799",
    "Santander Div.":  "C00000",
    "BB Dividendos":   "375623",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="1A1A2E", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def sv(ws, row, col, val):
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = val
    return cell

def fmt(ws, row, col, **kw):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    if "fill"   in kw: cell.fill          = kw["fill"]
    if "font"   in kw: cell.font          = kw["font"]
    if "align"  in kw: cell.alignment     = kw["align"]
    if "border" in kw: cell.border        = kw["border"]
    if "nf"     in kw: cell.number_format = kw["nf"]

# ═══════════════════════════════════════════════════════════════════
# 1. PREPROCESSAR CARTEIRAS (substituições + exclusões)
# ═══════════════════════════════════════════════════════════════════
def preprocessar_carteira(nome, pesos_brutos, colunas_csv):
    """Aplica substituições, remove tickers ausentes, renormaliza."""
    pesos = {}
    substituicoes_feitas = []
    excluidos = []

    for tk, w in pesos_brutos.items():
        tk_novo = SUBSTITUICOES.get(tk, tk)
        if tk_novo != tk:
            substituicoes_feitas.append((tk, tk_novo))
            tk = tk_novo
        if tk in colunas_csv:
            pesos[tk] = pesos.get(tk, 0.0) + w
        else:
            excluidos.append((tk, w))

    # Renormalizar
    total = sum(pesos.values())
    if total > 0:
        pesos = {k: v / total for k, v in pesos.items()}

    if substituicoes_feitas:
        print(f"  [{nome}] Substituições: {substituicoes_feitas}")
    if excluidos:
        print(f"  [{nome}] Excluídos (sem dados): {excluidos}")
        print(f"  [{nome}] Pesos renormalizados para 100%")

    return pesos

# ═══════════════════════════════════════════════════════════════════
# 2. CARREGAR PREÇOS
# ═══════════════════════════════════════════════════════════════════
print("─" * 60)
print("ANALISADOR DE CARTEIRAS DIVIDENDOS — Junho 2026")
print("─" * 60)
print(f"\nCarregando preços: {CSV_PATH}")
df_all = pd.read_csv(CSV_PATH, index_col=0, parse_dates=True).sort_index()
df_all = df_all.dropna(how="all")
colunas_csv = set(df_all.columns)
print(f"  {df_all.shape[0]} pregões | {df_all.shape[1]} ativos | "
      f"{df_all.index[0].date()} → {df_all.index[-1].date()}")

# ═══════════════════════════════════════════════════════════════════
# 3. PRÉ-PROCESSAR TODAS AS CARTEIRAS
# ═══════════════════════════════════════════════════════════════════
print("\nPré-processando carteiras...")
CARTEIRAS_PROC = {}
for nome, pesos in CARTEIRAS.items():
    CARTEIRAS_PROC[nome] = preprocessar_carteira(nome, pesos, colunas_csv)

# ─── Benchmark IBOV ───────────────────────────────────────────────
IBOV_SERIE = None
if "IBOV" in df_all.columns:
    IBOV_SERIE = df_all["IBOV"]

# ═══════════════════════════════════════════════════════════════════
# 4. FUNÇÕES CORE — MÉTRICAS E OTIMIZAÇÃO
# ═══════════════════════════════════════════════════════════════════
def calcular_log_rets(tickers, df_all):
    """Retorna array de log-retornos e série IBOV alinhada."""
    df = df_all[tickers].ffill(limit=5).dropna(how="all")
    prices = df.values.astype(float)
    log_rets = np.log(prices[1:] / prices[:-1])
    dates = df.index.tolist()

    if IBOV_SERIE is not None:
        ib = IBOV_SERIE.reindex(df.index).ffill(limit=5).values.astype(float)
        ibov_log = np.where(
            np.isnan(ib[1:]) | np.isnan(ib[:-1]),
            np.nan, np.log(ib[1:] / ib[:-1])
        )
    else:
        ibov_log = None

    return log_rets, ibov_log, dates


def risk_metrics(w, lr_clean, ibov_log=None, n_ret=None):
    """Calcula conjunto completo de métricas de risco para um vetor de pesos."""
    # Métricas anualizadas
    mu_d  = np.mean(lr_clean, axis=0)
    cov_d = np.cov(lr_clean.T, ddof=1)
    mu_a  = mu_d  * 252
    cov_a = cov_d * 252

    r_a  = float(w @ mu_a)
    vol  = float(np.sqrt(np.clip(w @ cov_a @ w, 0, None)))
    sharpe = (r_a - RF_AA) / vol if vol > 1e-12 else -999.0

    port_ret = lr_clean @ w

    # Sortino
    excess  = port_ret - RF_DIARIO
    neg_exc = excess[excess < 0]
    sd_down = math.sqrt(np.mean(neg_exc**2) * 252) if len(neg_exc) > 0 else 1e-12
    sortino = (r_a - RF_AA) / sd_down

    # VaR histórico
    var95_h = float(np.percentile(port_ret, 5))
    var99_h = float(np.percentile(port_ret, 1))

    # VaR paramétrico
    mu_p  = float(np.mean(port_ret))
    sig_p = float(np.std(port_ret, ddof=1))
    var95_p = mu_p - 1.6449 * sig_p
    var99_p = mu_p - 2.3263 * sig_p

    # CVaR
    cvar95 = float(np.mean(port_ret[port_ret <= var95_h]))
    cvar99 = float(np.mean(port_ret[port_ret <= var99_h]))

    # MDD
    cum  = np.exp(np.cumsum(port_ret))
    peak = np.maximum.accumulate(cum)
    mdd  = float(np.min((cum - peak) / peak))

    # Beta e Treynor
    if ibov_log is not None:
        ib_clean = np.where(np.isnan(ibov_log), 0.0, ibov_log)
        ib = ib_clean[:len(port_ret)]
        cov_pi = float(np.cov(port_ret, ib, ddof=1)[0, 1])
        var_ib = float(np.var(ib, ddof=1))
        beta   = cov_pi / var_ib if var_ib > 1e-14 else 1.0
    else:
        beta = 1.0
    treynor = (r_a - RF_AA) / beta if abs(beta) > 1e-12 else 0.0

    return {
        "R_anual":   r_a,
        "Vol_anual": vol,
        "Sharpe":    sharpe,
        "Sortino":   sortino,
        "Treynor":   treynor,
        "Beta":      beta,
        "VaR95_h":   var95_h,
        "VaR99_h":   var99_h,
        "VaR95_p":   var95_p,
        "VaR99_p":   var99_p,
        "CVaR95":    cvar95,
        "CVaR99":    cvar99,
        "MDD":       mdd,
    }


def otimizar_max_sharpe_customizado(w_rec, lr_clean, ibov_log=None):
    """
    Maximiza Sharpe com restrições por ativo:
        Σ wᵢ = 1
        α · wᵢ_rec ≤ wᵢ ≤ β · wᵢ_rec
        wᵢ ≥ ε > 0  (nenhum ativo zerado)

    Onde α = ALPHA (0.5) e β = BETA (2.0).
    """
    n = len(w_rec)
    EPS = 1e-4  # peso mínimo absoluto (garante nenhum ativo vai a zero)

    lb = np.maximum(ALPHA * w_rec, EPS)
    ub = np.minimum(BETA  * w_rec, 1.0)

    # Verificar viabilidade: lb deve ser ≤ ub para cada ativo
    # e sum(lb) deve ser ≤ 1
    if np.sum(lb) > 1.0 + 1e-6:
        # Escalar lb proporcionalmente
        lb = lb / np.sum(lb) * 0.99
    ub = np.maximum(ub, lb + 1e-5)

    mu_d  = np.mean(lr_clean, axis=0)
    cov_d = np.cov(lr_clean.T, ddof=1)
    mu_a  = mu_d * 252
    cov_a = cov_d * 252

    def neg_sharpe(w):
        r  = float(w @ mu_a)
        v  = float(np.sqrt(np.clip(w @ cov_a @ w, 0, None)))
        return -(r - RF_AA) / v if v > 1e-12 else 999.0

    constraints = [{"type": "eq", "fun": lambda w: w.sum() - 1.0}]
    bounds = Bounds(lb=lb, ub=ub)

    # Ponto inicial: pesos recomendados (já são factíveis por construção)
    w0 = np.clip(w_rec.copy(), lb, ub)
    w0 = w0 / w0.sum()

    res = minimize(neg_sharpe, w0, method="SLSQP",
                   bounds=bounds, constraints=constraints,
                   options={"maxiter": 5000, "ftol": 1e-14})

    w_opt = np.clip(res.x, lb, ub)
    w_opt = w_opt / w_opt.sum()
    return w_opt, res.success, res.fun


# ═══════════════════════════════════════════════════════════════════
# 5. RODAR ANÁLISE PARA CADA CARTEIRA
# ═══════════════════════════════════════════════════════════════════
print("\nCalculando métricas e otimizando...")

RESULTADOS = {}   # nome → dict com tudo

for nome, pesos_dict in CARTEIRAS_PROC.items():
    print(f"\n  ▶ {nome}")
    tickers = list(pesos_dict.keys())
    w_rec   = np.array([pesos_dict[t] for t in tickers])
    w_rec  /= w_rec.sum()   # garantir normalização

    # Dados de preço
    lr, ibov_log, dates = calcular_log_rets(tickers, df_all)
    lr_clean = np.where(np.isnan(lr), 0.0, lr)

    # Métricas como recomendado
    m_rec  = risk_metrics(w_rec, lr_clean, ibov_log)

    # Otimização Max Sharpe com bounds customizados
    w_opt, ok, obj_val = otimizar_max_sharpe_customizado(w_rec, lr_clean, ibov_log)
    m_opt  = risk_metrics(w_opt, lr_clean, ibov_log)

    print(f"    {len(tickers)} ativos | Sharpe rec={m_rec['Sharpe']:.3f} → opt={m_opt['Sharpe']:.3f} | "
          f"Solver {'OK' if ok else 'AVISO-não-convergiu'}")

    RESULTADOS[nome] = {
        "tickers": tickers,
        "w_rec":   w_rec,
        "w_opt":   w_opt,
        "m_rec":   m_rec,
        "m_opt":   m_opt,
        "dates":   dates,
        "ok":      ok,
    }


# ═══════════════════════════════════════════════════════════════════
# 6. GERAR EXCEL
# ═══════════════════════════════════════════════════════════════════
print(f"\nGerando Excel: {OUT_PATH}")
wb = openpyxl.Workbook()

# ──────────────────────────────────────────────────────────────────
# ABA 1: RESUMO COMPARATIVO
# ──────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Comparação"

nomes = list(RESULTADOS.keys())
N = len(nomes)

# Configurar larguras
ws.column_dimensions["A"].width = 28
for ci in range(N):
    for prefix in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        col_letter = get_column_letter(2 + ci * 2)
        ws.column_dimensions[col_letter].width = 12

for ci in range(N * 2 + 2):
    ws.column_dimensions[get_column_letter(2 + ci)].width = 13

R = 1

# ── Título ─────────────────────────────────────────────────────────
sv(ws, R, 1, "CARTEIRAS DE DIVIDENDOS · JUNHO 2026 · COMPARAÇÃO & OTIMIZAÇÃO MARKOWITZ (Mean-Variance)")
ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=1 + N * 2)
ws.row_dimensions[R].height = 24
fmt(ws, R, 1,
    fill=fill(C_NAVY), font=font(bold=True, color=C_WHITE, size=12),
    align=align("center"), border=thin())
R += 1

# ── Disclaimer ─────────────────────────────────────────────────────
disc = ("⚠ Análise retroativa (backward-looking) · NÃO constitui recomendação de investimento · "
        "Otimização Markowitz sujeita a maximização de erro de estimação · Dados: preços ajustados (snapshot estático)")
sv(ws, R, 1, disc)
ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=1 + N * 2)
ws.row_dimensions[R].height = 30
fmt(ws, R, 1,
    fill=fill(C_YELLOW), font=font(size=9, italic=True, color="C55A11", bold=True),
    align=align("center", wrap=True), border=thin())
R += 2

# ── Cabeçalhos das instituições ─────────────────────────────────────
sv(ws, R, 1, "Métrica")
fmt(ws, R, 1,
    fill=fill(C_NAVY), font=font(bold=True, color=C_WHITE),
    align=align("left"), border=thin())

for ci, nome in enumerate(nomes):
    col_rec = 2 + ci * 2
    col_opt = 3 + ci * 2
    cor = CORES_CARTEIRAS.get(nome, C_BLUE)

    sv(ws, R, col_rec, f"{nome}\nRecomendada")
    ws.merge_cells(start_row=R, start_column=col_rec, end_row=R, end_column=col_opt)
    ws.row_dimensions[R].height = 32
    fmt(ws, R, col_rec,
        fill=fill(cor), font=font(bold=True, color=C_WHITE, size=10),
        align=align("center", wrap=True), border=thin())
R += 1

# ── Sub-cabeçalhos Rec / Otim ──────────────────────────────────────
sv(ws, R, 1, "")
fmt(ws, R, 1, fill=fill(C_NAVY), border=thin())
for ci, nome in enumerate(nomes):
    col_rec = 2 + ci * 2
    col_opt = 3 + ci * 2
    cor = CORES_CARTEIRAS.get(nome, C_BLUE)

    sv(ws, R, col_rec, "Como Rec.")
    fmt(ws, R, col_rec,
        fill=fill(cor), font=font(bold=True, color=C_WHITE, size=9),
        align=align("center"), border=thin())

    sv(ws, R, col_opt, "Otimizada")
    fmt(ws, R, col_opt,
        fill=fill(C_NAVY), font=font(bold=True, color=C_WHITE, size=9),
        align=align("center"), border=thin())
R += 1

# ── Linhas de métricas ─────────────────────────────────────────────
METRICAS = [
    ("Nº de Ativos",         None,         "0",       None),
    ("Retorno Anual (log)",  "R_anual",    "0.00%",   "high"),
    ("Volatilidade Anual",   "Vol_anual",  "0.00%",   "low"),
    ("Sharpe Ratio",         "Sharpe",     "0.00",    "high"),
    ("Sortino Ratio",        "Sortino",    "0.00",    "high"),
    ("Treynor Ratio",        "Treynor",    "0.00%",   "high"),
    ("Beta (vs IBOV)",       "Beta",       "0.00",    None),
    ("VaR 95% Histórico",    "VaR95_h",   "0.000%",  "abs_low"),
    ("VaR 99% Histórico",    "VaR99_h",   "0.000%",  "abs_low"),
    ("CVaR 95% (ES)",        "CVaR95",    "0.000%",  "abs_low"),
    ("CVaR 99% (ES)",        "CVaR99",    "0.000%",  "abs_low"),
    ("Máximo Drawdown",      "MDD",       "0.00%",   "abs_low"),
]

for midx, (label, key, nf, best_rule) in enumerate(METRICAS):
    bg = C_GRAY_LT if midx % 2 == 0 else C_WHITE

    sv(ws, R, 1, label)
    fmt(ws, R, 1, fill=fill(bg), font=font(bold=True, size=10),
        align=align("left"), border=thin())

    # Coletar todos os valores para highlight
    all_vals_rec = []
    all_vals_opt = []
    for nome in nomes:
        res = RESULTADOS[nome]
        if key is None:
            all_vals_rec.append(len(res["tickers"]))
            all_vals_opt.append(len(res["tickers"]))
        else:
            all_vals_rec.append(res["m_rec"][key])
            all_vals_opt.append(res["m_opt"][key])

    # Melhor valor em cada coluna (rec vs opt separados)
    def best_val(vals, rule):
        if rule == "high":    return max(vals)
        if rule == "low":     return min(vals)
        if rule == "abs_low": return min(vals, key=abs)
        return None

    bv_rec = best_val(all_vals_rec, best_rule)
    bv_opt = best_val(all_vals_opt, best_rule)

    for ci, nome in enumerate(nomes):
        col_rec = 2 + ci * 2
        col_opt = 3 + ci * 2
        vr = all_vals_rec[ci]
        vo = all_vals_opt[ci]

        sv(ws, R, col_rec, vr)
        sv(ws, R, col_opt, vo)
        ws.cell(R, col_rec).number_format = nf
        ws.cell(R, col_opt).number_format = nf

        bg_rec = C_GREEN if (bv_rec is not None and abs(vr - bv_rec) < 1e-9) else bg
        bg_opt = C_GREEN if (bv_opt is not None and abs(vo - bv_opt) < 1e-9) else bg

        fmt(ws, R, col_rec, fill=fill(bg_rec), font=font(size=10),
            align=align("center"), border=thin())
        fmt(ws, R, col_opt, fill=fill(bg_opt), font=font(size=10),
            align=align("center"), border=thin())

    R += 1

R += 1

# ── Nota de parâmetros ─────────────────────────────────────────────
nota = (f"Parâmetros: RF = {RF_AA*100:.2f}% a.a. (CDI) | 252 pregões | log-retornos | long-only | "
        f"α = {ALPHA} | β = {BETA} | Max Sharpe com bounds por ativo | "
        f"Período: {df_all.index[0].date()} → {df_all.index[-1].date()}")
sv(ws, R, 1, nota)
ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=1 + N * 2)
fmt(ws, R, 1,
    fill=fill(C_GRAY_LT), font=font(size=9, italic=True, color="595959"),
    align=align("left"), border=thin())
R += 2

# ── Formulação OR ──────────────────────────────────────────────────
sv(ws, R, 1, "FORMULAÇÃO DO PROBLEMA DE OTIMIZAÇÃO (estilo Pesquisa Operacional)")
ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=1 + N * 2)
ws.row_dimensions[R].height = 22
fmt(ws, R, 1,
    fill=fill(C_TEAL), font=font(bold=True, color=C_WHITE, size=11),
    align=align("center"), border=thin())
R += 1

OR_LINES = [
    ("Variáveis de decisão:",  "wᵢ = peso do ativo i na carteira otimizada,  i = 1, …, n"),
    ("Dados de entrada:",      "μᵢ = retorno anual esperado do ativo i  |  σᵢⱼ = covariância anual entre ativos i e j  |  wᵢᵣₑc = peso recomendado pela instituição"),
    ("Função objetivo:",       "max   Sharpe(w) = [Rₚ(w) − Rᶠ] / σₚ(w)"),
    ("",                       "onde  Rₚ(w) = Σᵢ wᵢ μᵢ   (retorno esperado anualizado)"),
    ("",                       "      σₚ(w) = √(wᵀ Σ w)    (volatilidade anualizada)"),
    ("",                       "      Rᶠ = 14,40% a.a.      (CDI — taxa livre de risco)"),
    ("Restrições:",             "Σᵢ wᵢ = 1                  (pesos somam 100%)"),
    ("",                       f"α · wᵢᵣₑc ≤ wᵢ ≤ β · wᵢᵣₑc  para todo i  (bounds relativos; α = {ALPHA}, β = {BETA})"),
    ("",                       "wᵢ > 0  para todo i          (nenhum ativo zerado)"),
    ("Solver:",                 "scipy.optimize.minimize — SLSQP (Sequential Least Squares Programming)"),
    ("Observação:",            "Cada carteira define seu próprio conjunto de ativos e pesos recomendados. "
                               "A otimização é conduzida in-sample sobre os 252 pregões do snapshot. "
                               "Sharpe otimizado é um limite superior dentro da amostra."),
]

for lbl, txt in OR_LINES:
    sv(ws, R, 1, lbl)
    sv(ws, R, 2, txt)
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=1 + N * 2)
    lbl_bold = bool(lbl.strip())
    fmt(ws, R, 1,
        fill=fill(C_BLUE_LT), font=font(bold=lbl_bold, size=10, color=C_TEAL),
        align=align("left"), border=thin())
    fmt(ws, R, 2,
        fill=fill(C_GRAY_LT), font=font(size=10),
        align=align("left"), border=thin())
    R += 1


# ──────────────────────────────────────────────────────────────────
# ABA 2: PESOS POR CARTEIRA (detalhe)
# ──────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Pesos Detalhados")
ws2.column_dimensions["A"].width = 20

R2 = 1
sv(ws2, R2, 1, "PESOS POR ATIVO — RECOMENDADOS vs OTIMIZADOS")
ws2.merge_cells(start_row=R2, start_column=1, end_row=R2, end_column=5)
ws2.row_dimensions[R2].height = 22
fmt(ws2, R2, 1,
    fill=fill(C_NAVY), font=font(bold=True, color=C_WHITE, size=12),
    align=align("center"), border=thin())
R2 += 1

disc2 = f"Restrições de otimização: α · wᵢᵣₑc ≤ wᵢ ≤ β · wᵢᵣₑc  (α = {ALPHA}, β = {BETA}) | Nenhum ativo vai a zero"
sv(ws2, R2, 1, disc2)
ws2.merge_cells(start_row=R2, start_column=1, end_row=R2, end_column=5)
ws2.row_dimensions[R2].height = 20
fmt(ws2, R2, 1,
    fill=fill(C_YELLOW), font=font(size=9, italic=True, color=C_ORANGE, bold=True),
    align=align("center"), border=thin())
R2 += 2

for nome, res in RESULTADOS.items():
    cor = CORES_CARTEIRAS.get(nome, C_BLUE)
    tickers = res["tickers"]
    w_rec = res["w_rec"]
    w_opt = res["w_opt"]

    # Título da carteira
    sv(ws2, R2, 1, nome)
    ws2.merge_cells(start_row=R2, start_column=1, end_row=R2, end_column=5)
    ws2.row_dimensions[R2].height = 20
    fmt(ws2, R2, 1,
        fill=fill(cor), font=font(bold=True, color=C_WHITE, size=11),
        align=align("center"), border=thin())
    R2 += 1

    # Cabeçalho da tabela
    for ci, h in enumerate(["Ativo", "Peso Rec.", "Limite Inf (α)", "Limite Sup (β)", "Peso Otim.", "Δ Otim - Rec"]):
        sv(ws2, R2, ci+1, h)
        fmt(ws2, R2, ci+1,
            fill=fill(C_BLUE), font=font(bold=True, color=C_WHITE, size=9),
            align=align("center"), border=thin())
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 14
    R2 += 1

    for i, tk in enumerate(tickers):
        wr = float(w_rec[i])
        wo = float(w_opt[i])
        lb = ALPHA * wr
        ub = BETA  * wr
        d  = wo - wr
        bg = C_GRAY_LT if i % 2 == 0 else C_WHITE

        sv(ws2, R2, 1, tk)
        sv(ws2, R2, 2, wr)
        sv(ws2, R2, 3, lb)
        sv(ws2, R2, 4, ub)
        sv(ws2, R2, 5, wo)
        sv(ws2, R2, 6, d)

        for ci in range(1, 7):
            fmt(ws2, R2, ci,
                fill=fill(bg),
                font=font(size=10, bold=(ci==1)),
                align=align("left" if ci==1 else "center"),
                border=thin())
            if ci > 1:
                ws2.cell(R2, ci).number_format = "0.00%"

        # Delta: verde se sobe, vermelho se desce
        if   d >  0.001: fmt(ws2, R2, 6, fill=fill(C_GREEN))
        elif d < -0.001: fmt(ws2, R2, 6, fill=fill(C_RED_LT))
        R2 += 1

    R2 += 2


# ──────────────────────────────────────────────────────────────────
# ABA 3: DISCLAIMERS COMPLETOS
# ──────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Disclaimers")
ws3.column_dimensions["A"].width = 120
ws3.row_dimensions[1].height = 22

disc_full = [
    ("DISCLAIMERS E LIMITAÇÕES METODOLÓGICAS", True, C_NAVY),
    ("", False, C_WHITE),
    ("1. NATUREZA DA ANÁLISE", True, C_TEAL),
    ("Esta planilha é de natureza exclusivamente educacional e demonstrativa. Não constitui recomendação, oferta ou solicitação de compra ou venda de valores mobiliários. O objetivo é ilustrar conceitos de otimização de portfólio (Markowitz/Mean-Variance) e análise de risco quantitativo.", False, C_WHITE),
    ("", False, C_WHITE),
    ("2. DADOS HISTÓRICOS", True, C_TEAL),
    ("Toda a análise é retroativa (backward-looking). Os retornos, volatilidades, correlações e métricas de risco são calculados sobre preços ajustados históricos e refletem o passado, não o futuro. Desempenho passado não é garantia nem estimativa confiável de desempenho futuro.", False, C_WHITE),
    ("", False, C_WHITE),
    ("3. OTIMIZAÇÃO DE MARKOWITZ — LIMITAÇÕES CONHECIDAS", True, C_TEAL),
    ("a) Maximização de erro de estimação: a otimização amplifica os erros de estimação dos parâmetros de entrada (retornos esperados, variâncias e covariâncias). Os pesos 'ótimos' são instáveis e sensíveis a pequenas variações nos dados.", False, C_WHITE),
    ("b) O Sharpe otimizado é um limite superior in-sample. Aplicado a dados futuros (out-of-sample), o Sharpe tipicamente degrada de forma substancial.", False, C_WHITE),
    ("c) A janela histórica utilizada (252 pregões) é curta para estimação robusta da matriz de covariância. Com n ativos próximos do número de observações, a matriz amostral é instável.", False, C_WHITE),
    ("d) Correlações não são estacionárias. Em períodos de crise, correlações tendem a subir em direção a 1, justamente quando a diversificação é mais necessária.", False, C_WHITE),
    ("", False, C_WHITE),
    ("4. RESTRIÇÕES DE PESO", True, C_TEAL),
    (f"Os limites por ativo (α = {ALPHA}, β = {BETA}) foram definidos pelo usuário para manter a carteira otimizada próxima à composição institucional recomendada. Esses limites são arbitrários e não derivam de análise quantitativa.", False, C_WHITE),
    ("", False, C_WHITE),
    ("5. SUBSTITUIÇÕES DE TICKERS", True, C_TEAL),
    ("AXIA6 (PN Axia Energia, ex-ELET6) substituído por AXIA3 (ON) por ausência de histórico no snapshot. As duas classes são altamente correlacionadas mas não idênticas. ROXO34 (BDR Nubank) excluído do portfólio Casa A por ausência de dados; os pesos foram renormalizados.", False, C_WHITE),
    ("", False, C_WHITE),
    ("6. CONFORMIDADE REGULATÓRIA", True, C_TEAL),
    ("Analistas e assessores de investimento regulados pela CVM devem observar as Resoluções CVM 20 e afins antes de compartilhar análises de carteiras com clientes. Esta análise não foi produzida por analista certificado e não substitui parecer de compliance ou jurídico.", False, C_WHITE),
    ("", False, C_WHITE),
    (f"Parâmetros utilizados: RF = {RF_AA*100:.2f}% a.a. | α = {ALPHA} | β = {BETA} | Período: {df_all.index[0].date()} → {df_all.index[-1].date()}", False, C_GRAY_LT),
]

for ridx, (txt, bold, bg) in enumerate(disc_full, start=1):
    sv(ws3, ridx, 1, txt)
    fmt(ws3, ridx, 1,
        fill=fill(bg),
        font=font(bold=bold, color=C_WHITE if bg != C_WHITE else "1A1A2E", size=10 if bold else 9),
        align=align("left"), border=thin())
    if not txt:
        ws3.row_dimensions[ridx].height = 8
    elif bold:
        ws3.row_dimensions[ridx].height = 20
    else:
        ws3.row_dimensions[ridx].height = 30
        fmt(ws3, ridx, 1, align=align("left", wrap=True))


# ══════════════════════════════════════════════════════════════════
# SALVAR
# ══════════════════════════════════════════════════════════════════
wb.save(OUT_PATH)
print(f"\n✓ Arquivo salvo: {OUT_PATH}")

# ═══════════════════════════════════════════════════════════════════
# RESUMO NO TERMINAL
# ═══════════════════════════════════════════════════════════════════
print(f"\n{'═'*78}")
print(f"  {'CARTEIRA':<22} {'SHARPE REC':>12}  {'SHARPE OTM':>12}  {'DELTA':>10}  {'VOL REC':>10}")
print(f"{'─'*78}")
for nome, res in RESULTADOS.items():
    sr = res['m_rec']['Sharpe']
    so = res['m_opt']['Sharpe']
    vr = res['m_rec']['Vol_anual']
    print(f"  {nome:<22} {sr:>12.3f}  {so:>12.3f}  {so-sr:>+10.3f}  {vr:>10.1%}")
print(f"{'═'*78}")
print(f"\nFormulação OR: Σ wᵢ = 1 | α·wᵢᵣₑc ≤ wᵢ ≤ β·wᵢᵣₑc | α={ALPHA}, β={BETA}")
print(f"Objetivo: max Sharpe = (Rₚ - Rᶠ) / σₚ  |  Rᶠ = {RF_AA:.2%} a.a.")
print(f"\n⚠  Análise retroativa. Não é recomendação de investimento.")
print(f"   Desempenho passado não garante desempenho futuro.")

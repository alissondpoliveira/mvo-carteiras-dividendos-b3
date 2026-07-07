"""
atualizar_carteira.py — Atualiza minha_carteira.xlsx com valores calculados em Python.

Uso: python atualizar_carteira.py
     (Execute sempre que mudar tickers ou quantidades em Minha Carteira)
"""
import openpyxl
import pandas as pd
import numpy as np
import math
import os

BASE = os.path.dirname(os.path.abspath(__file__))
XLS  = os.path.join(BASE, "..", "planilhas", "minha_carteira.xlsx")
CSV  = os.path.join(BASE, "..", "dados", "precos_b3.csv")

RF_AA  = 0.1440   # CDI ~14.40% a.a. — atualize conforme necessário
RF_D   = math.log(1 + RF_AA) / 252
N_ROWS = 500

# ── 1. Ler tickers e quantidades ─────────────────────────────────────────────
wb = openpyxl.load_workbook(XLS)
ws_mc = wb["Minha Carteira"]

tickers, qtds = [], []
for r in range(6, 26):
    t = ws_mc.cell(row=r, column=1).value
    q = ws_mc.cell(row=r, column=2).value
    if t and q and isinstance(q, (int, float)):
        tickers.append(str(t).strip())
        qtds.append(float(q))

if not tickers:
    print("Nenhum ativo em Minha Carteira. Adicione ticker + quantidade.")
    exit(1)

print(f"Ativos: {tickers}")

# ── 2. Carregar preços ────────────────────────────────────────────────────────
df = pd.read_csv(CSV, index_col=0)
missing = [t for t in tickers if t not in df.columns]
if missing:
    print(f"Tickers não encontrados: {missing}")
    print(f"Disponíveis: {list(df.columns[:20])}")
    exit(1)

# ── 3. Calcular retornos e pesos ──────────────────────────────────────────────
prices   = df[tickers + ["IBOV", "DIVO11"]].copy()
log_rets = np.log(prices / prices.shift(1)).dropna()

last_prices = prices.iloc[-1][tickers]
valores     = [qtds[i] * last_prices[tickers[i]] for i in range(len(tickers))]
total       = sum(valores)
pesos       = [v / total for v in valores]

rets_port = (log_rets[tickers] * pesos).sum(axis=1)
rets_ibov = log_rets["IBOV"]
rets_divo = log_rets["DIVO11"]
n         = len(rets_port)

print(f"Pregões: {n}  |  {log_rets.index[0]} → {log_rets.index[-1]}")

cum_idx  = np.exp(np.cumsum(rets_port.values))
peak     = np.maximum.accumulate(cum_idx)
drawdown = cum_idx / peak - 1

# ── 4. Gravar _Calc com valores estáticos ─────────────────────────────────────
ws_calc = wb["_Calc"]

for r in range(2, N_ROWS + 2):
    for c in range(1, 26):
        ws_calc.cell(row=r, column=c).value = None

headers = [f"Ret_{i+1}" for i in range(20)] + ["Ret_IBOV","Ret_DIVO11","Ret_Port","Idx_Cum","Drawdown"]
for c, h in enumerate(headers, 1):
    ws_calc.cell(row=1, column=c).value = h

for i, ticker in enumerate(tickers):
    for r_idx, val in enumerate(log_rets[ticker].values, start=2):
        ws_calc.cell(row=r_idx, column=i+1).value = float(val)

for r_idx, (wp, wi, wd, ci, dd) in enumerate(
    zip(rets_port, rets_ibov, rets_divo, cum_idx, drawdown), start=2):
    ws_calc.cell(row=r_idx, column=21).value = float(wi)
    ws_calc.cell(row=r_idx, column=22).value = float(wd)
    ws_calc.cell(row=r_idx, column=23).value = float(wp)
    ws_calc.cell(row=r_idx, column=24).value = float(ci)
    ws_calc.cell(row=r_idx, column=25).value = float(dd)

# ── 5. Atualizar Resultados ───────────────────────────────────────────────────
ws_res = wb["Resultados"]
W  = f"_Calc!$W$2:$W${n+1}"
U  = f"_Calc!$U$2:$U${n+1}"
V  = f"_Calc!$V$2:$V${n+1}"
Y  = f"_Calc!$Y$2:$Y${n+1}"

ws_res.cell(row=6,  column=2).value = RF_AA
ws_res.cell(row=7,  column=2).value = "=LN(1+B6)/252"
ws_res.cell(row=8,  column=2).value = n
ws_res.cell(row=9,  column=2).value = len(tickers)
ws_res.cell(row=10, column=2).value = round(total, 2)

for col, rng in [(2, W), (3, U), (4, V)]:
    ws_res.cell(row=15, column=col).value = f"=AVERAGE({rng})*252"
    ws_res.cell(row=16, column=col).value = f"=STDEV({rng})*SQRT(252)"

ws_res.cell(row=17, column=2).value = f"={W.split(':')[1]}-1"

for col in [2, 3, 4]:
    c = chr(64 + col)
    ws_res.cell(row=19, column=col).value = f"=IFERROR(({c}15-$B$6)/{c}16,\"N/D\")"

N = str(n)
ws_res.cell(row=20, column=2).value = (
    f"=IFERROR((B15-$B$6)/SQRT(SUMPRODUCT((({W}-$B$7)*({W}<$B$7))^2)/{N}*252),\"N/D\")")
ws_res.cell(row=20, column=3).value = (
    f"=IFERROR((C15-$B$6)/SQRT(SUMPRODUCT((({U}-$B$7)*({U}<$B$7))^2)/{N}*252),\"N/D\")")

for col, rng in [(2, W), (3, U), (4, V)]:
    c = chr(64 + col)
    ws_res.cell(row=22, column=col).value = f"=PERCENTILE({rng},0.05)"
    ws_res.cell(row=23, column=col).value = f"=AVERAGEIF({rng},\"<=\"&{c}22)"
    ws_res.cell(row=24, column=col).value = f"=PERCENTILE({rng},0.01)"
    ws_res.cell(row=25, column=col).value = f"=AVERAGEIF({rng},\"<=\"&{c}24)"

ws_res.cell(row=26, column=2).value = f"=MIN({Y})"
ws_res.cell(row=28, column=2).value = f"=IFERROR(COVARIANCE.S({W},{U})/VAR.S({U}),\"N/D\")"
ws_res.cell(row=29, column=2).value = "=IFERROR(B15-($B$6+B28*(C15-$B$6)),\"N/D\")"
ws_res.cell(row=30, column=2).value = "=IFERROR((B15-$B$6)/B28,\"N/D\")"
ws_res.cell(row=31, column=2).value = (
    f"=IFERROR((AVERAGE({W})-AVERAGE({V}))*252/(STDEV({W}-{V})*SQRT(252)),\"N/D\")")

# ── 6. Atualizar preços em Minha Carteira ────────────────────────────────────
for r_idx, ticker in enumerate(tickers, start=6):
    ws_mc.cell(row=r_idx, column=3).value = round(float(last_prices[ticker]), 2)
    ws_mc.cell(row=r_idx, column=4).value = round(float(valores[r_idx-6]), 2)
    ws_mc.cell(row=r_idx, column=5).value = round(float(pesos[r_idx-6]), 6)

# ── 7. Salvar ─────────────────────────────────────────────────────────────────
wb.save(XLS)
print(f"\n✓ {XLS} atualizado")

ret_a   = rets_port.mean() * 252
vol_a   = rets_port.std()  * np.sqrt(252)
beta    = np.cov(rets_port, rets_ibov)[0,1] / np.var(rets_ibov, ddof=1)
print(f"  Retorno Anual: {ret_a*100:.2f}%  |  Volatilidade: {vol_a*100:.2f}%")
print(f"  Sharpe: {(ret_a-RF_AA)/vol_a:.3f}  |  Beta: {beta:.3f}")

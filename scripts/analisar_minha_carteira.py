# -*- coding: utf-8 -*-
"""
analisar_minha_carteira.py
══════════════════════════════════════════════════════════════════════
Preencha minha_carteira.xlsx com seus tickers e quantidade de papéis.
O script calcula os pesos pela cotação atual, executa a análise
completa de risco-retorno e gera um dashboard em resultado_carteira.xlsx

Uso:
    python scripts/analisar_minha_carteira.py

Saída:
    resultado_carteira.xlsx  (pasta raiz do projeto)

══════════════════════════════════════════════════════════════════════
DISCLAIMERS:
  • Análise retroativa. NÃO é recomendação de investimento.
  • O autor não é analista credenciado pela CVM.
  • Rentabilidade passada não garante resultados futuros.
══════════════════════════════════════════════════════════════════════
"""

import os, math, warnings
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from scipy.optimize import minimize, Bounds

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════
# CONFIGURAÇÕES
# ══════════════════════════════════════════════════════════════════════
PASTA      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PRECOS = os.path.join(PASTA, "dados",     "precos_b3.csv")
ENTRADA    = os.path.join(PASTA, "planilhas", "minha_carteira.xlsx")
SAIDA      = os.path.join(PASTA,              "resultado_carteira.xlsx")

RF_AA  = 0.1440   # taxa livre de risco anual (Selic atual)
ALPHA  = 0.5      # limite inferior = 50 % do peso original
BETA   = 2.0      # limite superior = 200 % do peso original

BENCHMARK_IBOV = "IBOV"
BENCHMARK_DIVO = "DIVO11"

# ══════════════════════════════════════════════════════════════════════
# ESTILOS
# ══════════════════════════════════════════════════════════════════════
def fill(h):   return PatternFill("solid", fgColor=h)
def fnt(b=False, c="1A1A2E", s=10, i=False):
    return Font(bold=b, color=c, size=s, italic=i, name="Calibri")
def aln(h="center", w=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=w)
def brd(c="BFBFBF"):
    sd = Side(style="thin", color=c)
    return Border(left=sd, right=sd, top=sd, bottom=sd)

def cel(ws, r, c, v=None, **kw):
    cell = ws.cell(r, c)
    if isinstance(cell, MergedCell):
        return cell
    if v is not None:
        cell.value = v
    for k in ("fl","fn","al","br","nf"):
        if k not in kw: continue
        setattr(cell, {"fl":"fill","fn":"font","al":"alignment",
                       "br":"border","nf":"number_format"}[k], kw[k])
    return cell

# paleta
C1="1F3864"; C2="FFFFFF"; C3="FFF2CC"; C4="C6EFCE"
C5="FCE4D6"; C6="F5F5F5"; CO="C55A11"
CA="17375E"; CB="7030A0"

# ══════════════════════════════════════════════════════════════════════
# MÉTRICAS
# ══════════════════════════════════════════════════════════════════════
RF_D = math.log(1 + RF_AA) / 252

def metricas(w, lr, ibov_lr=None, divo_lr=None):
    lr    = np.nan_to_num(lr)
    mu_a  = np.mean(lr, 0) * 252
    cov_a = np.atleast_2d(np.cov(lr.T, ddof=1) * 252)
    Ra    = float(w @ mu_a)
    vol   = float(np.sqrt(np.clip(w @ cov_a @ w, 0, None)))
    sharpe  = (Ra - RF_AA) / vol if vol > 1e-12 else -999.0
    pr      = lr @ w
    excess  = pr - RF_D
    neg     = excess[excess < 0]
    sd_dn   = math.sqrt(np.mean(neg**2) * 252) if len(neg) else 1e-12
    sortino = (Ra - RF_AA) / sd_dn
    var95   = float(np.percentile(pr,  5))
    var99   = float(np.percentile(pr,  1))
    cvar95  = float(np.mean(pr[pr <= var95]))
    cvar99  = float(np.mean(pr[pr <= var99]))
    cum     = np.exp(np.cumsum(pr))
    peak    = np.maximum.accumulate(cum)
    mdd     = float(np.min((cum - peak) / peak))
    cum_r   = float(cum[-1] - 1)
    vol_a   = float(np.std(pr, ddof=1) * math.sqrt(252))

    beta = alpha_j = treynor = ir_divo = ra_ibov = None
    if ibov_lr is not None:
        ib      = np.nan_to_num(ibov_lr[:len(pr)])
        var_ib  = float(np.var(ib, ddof=1))
        cov_pi  = float(np.cov(pr, ib, ddof=1)[0, 1])
        beta    = cov_pi / var_ib if var_ib > 1e-14 else 1.0
        ra_ibov = float(np.mean(ib) * 252)
        alpha_j = Ra - (RF_AA + beta * (ra_ibov - RF_AA))
        treynor = (Ra - RF_AA) / beta if abs(beta) > 1e-12 else 0.0
    if divo_lr is not None:
        dv      = np.nan_to_num(divo_lr[:len(pr)])
        active  = pr - dv
        te      = float(np.std(active, ddof=1) * math.sqrt(252))
        ir_divo = float(np.mean(active) * 252 / te) if te > 1e-12 else 0.0

    return dict(Ra=Ra, vol=vol_a, sharpe=sharpe, sortino=sortino,
                var95=var95, var99=var99, cvar95=cvar95, cvar99=cvar99,
                mdd=mdd, cum_ret=cum_r, beta=beta, alpha_j=alpha_j,
                treynor=treynor, ir_divo=ir_divo, ra_ibov=ra_ibov)

def otimizar(w_rec, lr):
    lr  = np.nan_to_num(lr)
    EPS = 1e-4
    lb  = np.maximum(ALPHA * w_rec, EPS)
    ub  = np.minimum(BETA  * w_rec, 1.0)
    if np.sum(lb) > 1.0 + 1e-6:
        lb = lb / np.sum(lb) * 0.99
    ub  = np.maximum(ub, lb + 1e-5)
    mu_a  = np.mean(lr, 0) * 252
    cov_a = np.atleast_2d(np.cov(lr.T, ddof=1) * 252)
    def neg_sh(w):
        r = float(w @ mu_a)
        v = float(np.sqrt(np.clip(w @ cov_a @ w, 0, None)))
        return -(r - RF_AA) / v if v > 1e-12 else 999.0
    w0  = np.clip(w_rec.copy(), lb, ub); w0 /= w0.sum()
    res = minimize(neg_sh, w0, method="SLSQP",
                   bounds=Bounds(lb=lb, ub=ub),
                   constraints=[{"type": "eq", "fun": lambda w: w.sum() - 1.0}],
                   options={"maxiter": 5000, "ftol": 1e-14})
    w_opt = np.clip(res.x, lb, ub); w_opt /= w_opt.sum()
    return w_opt, res.success

# ══════════════════════════════════════════════════════════════════════
# LEITURA DO TEMPLATE
# ══════════════════════════════════════════════════════════════════════
print("=" * 62)
print("  Analisador de Carteira  |  Otimizador de Máximo Sharpe")
print("=" * 62)

for path, label in [(ENTRADA,"minha_carteira.xlsx"), (CSV_PRECOS,"precos_b3.csv")]:
    if not os.path.exists(path):
        print(f"\n[ERRO] Nao encontrado: {path}")
        exit(1)

wb_in = openpyxl.load_workbook(ENTRADA)
ws_in = wb_in["Minha Carteira"]

portfolio_raw = {}   # ticker -> quantidade de papeis
nome_carteira = "Minha Carteira"

for row in ws_in.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    ticker = str(row[0]).strip().upper()
    if ticker == "NOME_CARTEIRA":
        nome_carteira = str(row[1]).strip() if row[1] else nome_carteira
        continue
    if row[1] is not None:
        try:
            qtd = float(row[1])
            if qtd > 0:
                portfolio_raw[ticker] = qtd
        except (ValueError, TypeError):
            pass

if not portfolio_raw:
    print("[ERRO] Nenhum ativo encontrado em minha_carteira.xlsx")
    exit(1)

# ══════════════════════════════════════════════════════════════════════
# DADOS DE PREÇOS
# ══════════════════════════════════════════════════════════════════════
print(f"\nLendo precos_b3.csv...")
df_all = (pd.read_csv(CSV_PRECOS, index_col=0, parse_dates=True)
            .sort_index().dropna(how="all"))
colunas = df_all.columns.tolist()

# Verificar tickers e calcular pesos por valor financeiro
ausentes = [tk for tk in portfolio_raw if tk not in colunas]
if ausentes:
    print(f"  [AVISO] Nao encontrados no CSV (ignorados): {ausentes}")
    for tk in ausentes:
        del portfolio_raw[tk]
    if not portfolio_raw:
        print("[ERRO] Nenhum ativo valido apos remocao.")
        exit(1)

tickers   = list(portfolio_raw.keys())
quantidades = np.array([portfolio_raw[tk] for tk in tickers], dtype=float)

# Ultimo preco disponivel para cada ativo
ultimo_preco = np.array([df_all[tk].dropna().iloc[-1] for tk in tickers], dtype=float)
valores      = quantidades * ultimo_preco
valor_total  = valores.sum()
w_rec        = valores / valor_total

data_referencia = df_all.index[-1].date()

print(f"\n  Carteira  : {nome_carteira}")
print(f"  Ativos    : {len(tickers)}")
print(f"  Ref. pesos: {data_referencia} (ultima cotacao disponivel)")
print(f"  Valor est.: R$ {valor_total:,.2f}")
print()
for tk, qty, preco, val, w in zip(tickers, quantidades, ultimo_preco, valores, w_rec):
    print(f"  {tk:<8} {int(qty):>6} papeis  x  R${preco:>8.2f}  =  R${val:>10,.2f}  ({w:.1%})")

# ══════════════════════════════════════════════════════════════════════
# LOG-RETORNOS
# ══════════════════════════════════════════════════════════════════════
df_port = df_all[tickers].ffill(limit=5)
arr     = df_port.values.astype(float)
lr_port = np.where(np.isnan(arr[1:]) | np.isnan(arr[:-1]), 0.0,
                   np.log(np.clip(arr[1:] / arr[:-1], 1e-8, None)))

def lr_serie(col):
    if col not in df_all.columns: return None
    s = df_all[col].ffill(limit=5).values.astype(float)
    return np.where(np.isnan(s[1:]) | np.isnan(s[:-1]), 0.0,
                    np.log(np.clip(s[1:] / s[:-1], 1e-8, None)))

ibov_lr = lr_serie(BENCHMARK_IBOV)
divo_lr = lr_serie(BENCHMARK_DIVO)
inicio  = df_all.index[0].date()
fim     = df_all.index[-1].date()
n_dias  = len(lr_port)

# ══════════════════════════════════════════════════════════════════════
# ANÁLISE
# ══════════════════════════════════════════════════════════════════════
print("\nCalculando metricas de risco-retorno...")
m_rec  = metricas(w_rec,  lr_port, ibov_lr, divo_lr)

print("Otimizando pesos (Max Sharpe — SLSQP)...")
w_opt, ok = otimizar(w_rec, lr_port)
m_opt     = metricas(w_opt, lr_port, ibov_lr, divo_lr)

m_ibov = (metricas(np.array([1.0]), ibov_lr.reshape(-1,1), ibov_lr, divo_lr)
          if ibov_lr is not None else None)
m_divo = (metricas(np.array([1.0]), divo_lr.reshape(-1,1), ibov_lr, divo_lr)
          if divo_lr is not None else None)

# ══════════════════════════════════════════════════════════════════════
# EXCEL DASHBOARD
# ══════════════════════════════════════════════════════════════════════
print(f"\nGerando dashboard: {SAIDA}")
wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
# ABA 1: DASHBOARD
# ─────────────────────────────────────────────
ws = wb.active
ws.title = "Dashboard"

ws.column_dimensions["A"].width = 30
for c in ["B","C","D","E"]: ws.column_dimensions[c].width = 16
ws.column_dimensions["F"].width = 2  # spacer

R = 1

def h1(ws, r, text, span="A:E"):
    ws.merge_cells(f"{span[0]}{r}:{span[-1]}{r}")
    cel(ws, r, 1, text,
        fl=fill(C1), fn=fnt(b=True, c=C2, s=13), al=aln("left"), br=brd())
    ws.row_dimensions[r].height = 26

def h2(ws, r, text, color="1E6B6B"):
    ws.merge_cells(f"A{r}:E{r}")
    cel(ws, r, 1, text,
        fl=fill(color), fn=fnt(b=True, c=C2, s=9), al=aln("center"), br=brd())
    ws.row_dimensions[r].height = 15

# Titulo
h1(ws, R, f"  {nome_carteira}  —  Análise de Risco e Retorno")
R += 1
ws.merge_cells(f"A{R}:E{R}")
cel(ws, R, 1,
    f"Período de análise: {inicio} → {fim}  |  {n_dias} pregões  "
    f"|  RF = {RF_AA:.2%} a.a.  |  Data de referência dos pesos: {data_referencia}",
    fl=fill(C3), fn=fnt(c=CO, s=9, i=True), al=aln("left"), br=brd())
ws.row_dimensions[R].height = 16; R += 1

ws.merge_cells(f"A{R}:E{R}")
cel(ws, R, 1,
    "⚠ Análise retroativa. NÃO constitui recomendação de investimento. "
    "Rentabilidade passada não garante resultados futuros. "
    "O autor não é analista credenciado pela CVM.",
    fl=fill(C5), fn=fnt(c="C00000", s=8, i=True), al=aln("left", w=True), br=brd())
ws.row_dimensions[R].height = 20; R += 2

# ── Composição da Carteira ──
h2(ws, R, "  COMPOSIÇÃO DA CARTEIRA  (pesos calculados pela cotação de referência)")
R += 1
for ci, h in enumerate(["Ativo","Quantidade","Último Preço","Valor Financeiro","Peso %"], 1):
    cel(ws, R, ci, h,
        fl=fill("2E75B6"), fn=fnt(b=True, c=C2, s=10), al=aln("center"), br=brd())
ws.row_dimensions[R].height = 18; R += 1

for i, (tk, qty, preco, val, w) in enumerate(
        zip(tickers, quantidades, ultimo_preco, valores, w_rec)):
    bg = C6 if i % 2 == 0 else "FFFFFF"
    cel(ws, R, 1, tk,                fl=fill(bg), fn=fnt(b=True,s=10), al=aln("left"),   br=brd())
    cel(ws, R, 2, int(qty),          fl=fill(bg), fn=fnt(s=10),        al=aln("center"), br=brd(), nf="#,##0")
    cel(ws, R, 3, preco,             fl=fill(bg), fn=fnt(s=10),        al=aln("center"), br=brd(), nf='R$ #,##0.00')
    cel(ws, R, 4, val,               fl=fill(bg), fn=fnt(s=10),        al=aln("center"), br=brd(), nf='R$ #,##0.00')
    cel(ws, R, 5, w,                 fl=fill(bg), fn=fnt(s=10),        al=aln("center"), br=brd(), nf="0.00%")
    ws.row_dimensions[R].height = 18; R += 1

# Total
cel(ws, R, 1, "TOTAL",
    fl=fill("2E75B6"), fn=fnt(b=True,c=C2,s=10), al=aln("left"), br=brd())
cel(ws, R, 4, valor_total,
    fl=fill("2E75B6"), fn=fnt(b=True,c=C2,s=10), al=aln("center"), br=brd(), nf='R$ #,##0.00')
cel(ws, R, 5, 1.0,
    fl=fill("2E75B6"), fn=fnt(b=True,c=C2,s=10), al=aln("center"), br=brd(), nf="0.00%")
for ci in [2,3]:
    cel(ws, R, ci, fl=fill("2E75B6"), br=brd())
ws.row_dimensions[R].height = 18; R += 2

# ── Indicadores Risco-Retorno ──
for ci, (h, col) in enumerate(zip(
        ["Métrica", "Carteira (atual)", "Carteira Otimizada", "IBOV", "DIVO11"],
        [C1, "505050", CA, "404040", CB]), 1):
    cel(ws, R, ci, h,
        fl=fill(col), fn=fnt(b=True, c=C2, s=10), al=aln("center"), br=brd())
ws.row_dimensions[R].height = 18; R += 1

GRUPOS = [
    ("── RETORNO E RISCO ──", None, None, None, None, "sep"),
    ("Retorno Anual",         "Ra",      "0.00%",  "0.00%",  "0.00%",  None),
    ("Volatilidade Anual",    "vol",     "0.00%",  "0.00%",  "0.00%",  None),
    ("Retorno Acumulado",     "cum_ret", "0.00%",  "0.00%",  "0.00%",  None),
    ("── RISCO-RETORNO ──", None, None, None, None, "sep"),
    ("Índice de Sharpe",      "sharpe",  "+0.000", "+0.000", "+0.000", None),
    ("Índice de Sortino",     "sortino", "+0.000", "+0.000", "+0.000", None),
    ("Índice de Treynor",     "treynor", "+0.000", "+0.000", "+0.000", None),
    ("── RISCO DE CAUDA ──", None, None, None, None, "sep"),
    ("VaR 95% (diário)",     "var95",   "0.000%", "0.000%", "0.000%", None),
    ("CVaR 95% (diário)",    "cvar95",  "0.000%", "0.000%", "0.000%", None),
    ("VaR 99% (diário)",     "var99",   "0.000%", "0.000%", "0.000%", None),
    ("CVaR 99% (diário)",    "cvar99",  "0.000%", "0.000%", "0.000%", None),
    ("Maximum Drawdown",     "mdd",     "0.00%",  "0.00%",  "0.00%",  None),
    ("── BENCHMARK ──", None, None, None, None, "sep"),
    ("Beta vs IBOV",          "beta",    "+0.000", "+0.000", "+0.000", None),
    ("Alpha de Jensen",       "alpha_j", "+0.00%", "+0.00%", "+0.00%", None),
    ("IR vs DIVO11",          "ir_divo", "+0.000", "+0.000", "+0.000", None),
]

UP_BETTER  = {"Ra","sharpe","sortino","treynor","cum_ret","ir_divo","alpha_j"}
DOWN_BETTER = {"var95","var99","cvar95","cvar99","mdd","vol","beta"}

for idx, linha in enumerate(GRUPOS):
    if linha[5] == "sep":
        h2(ws, R, f"  {linha[0]}  ")
        R += 1; continue

    label, key, nf_rec, nf_opt, nf_bm = linha[0], linha[1], linha[2], linha[3], linha[4]
    bg = C6 if idx % 2 == 0 else "FFFFFF"

    cel(ws, R, 1, label, fl=fill(bg), fn=fnt(b=True, s=10), al=aln("left"), br=brd())

    v_rec  = m_rec.get(key)  if key else None
    v_opt  = m_opt.get(key)  if key else None
    v_ibov = m_ibov.get(key) if (key and m_ibov) else None
    v_divo = m_divo.get(key) if (key and m_divo) else None
    vals   = [v_rec, v_opt, v_ibov, v_divo]
    nfs    = [nf_rec, nf_opt, nf_bm, nf_bm]

    # highlight melhor entre rec vs opt
    pair = [v for v in [v_rec, v_opt] if v is not None]
    if pair and key:
        if key in UP_BETTER:
            best = max(pair)
        else:
            best = min(pair, key=abs)
    else:
        best = None

    for ci, (v, nf) in enumerate(zip(vals, nfs)):
        is_best = (best is not None and v is not None and ci < 2
                   and abs(v - best) < 1e-9)
        bg_c = C4 if is_best else bg
        cel(ws, R, 2+ci, v if v is not None else "—",
            fl=fill(bg_c),
            fn=fnt(s=10, c=("404040" if ci >= 2 else "1A1A2E")),
            al=aln("center"), br=brd(),
            nf=nf if v is not None else "@")
    ws.row_dimensions[R].height = 18; R += 1

# ─────────────────────────────────────────────
# ABA 2: PESOS OTIMIZADOS
# ─────────────────────────────────────────────
ws2 = wb.create_sheet("Pesos Otimizados")
ws2.column_dimensions["A"].width = 14
for c in ["B","C","D","E"]: ws2.column_dimensions[c].width = 16

R2 = 1
ws2.merge_cells("A1:E1")
cel(ws2,1,1, "Pesos: Carteira Atual vs. Otimizado (Max Sharpe)",
    fl=fill(C1), fn=fnt(b=True,c=C2,s=12), al=aln("center"), br=brd())
ws2.row_dimensions[1].height = 22; R2 = 3

for ci, (h, col) in enumerate(zip(
        ["Ativo","Peso Atual","Peso Otimizado","Δ Peso","Sugestão"],
        ["2E75B6","505050",CA,"404040","404040"]), 1):
    cel(ws2,R2,ci,h, fl=fill(col), fn=fnt(b=True,c=C2,s=10), al=aln("center"), br=brd())
ws2.row_dimensions[R2].height = 18; R2 += 1

for i, (tk, wr) in enumerate(zip(tickers, w_rec)):
    wo = float(w_opt[i]); delta = wo - wr
    bg = C6 if i % 2 == 0 else "FFFFFF"
    sug = ("▲ Aumentar" if delta > 0.005
           else "▼ Reduzir" if delta < -0.005
           else "≈ Manter")
    sug_c = "375623" if delta > 0.005 else ("843C0C" if delta < -0.005 else "404040")
    cel(ws2,R2,1,tk,    fl=fill(bg),fn=fnt(b=True,s=10),al=aln("left"),  br=brd())
    cel(ws2,R2,2,wr,    fl=fill(bg),fn=fnt(s=10),       al=aln("center"),br=brd(),nf="0.00%")
    cel(ws2,R2,3,wo,    fl=fill(bg),fn=fnt(s=10),       al=aln("center"),br=brd(),nf="0.00%")
    cel(ws2,R2,4,delta,
        fl=fill(C4 if delta>0.005 else (C5 if delta<-0.005 else bg)),
        fn=fnt(s=10),al=aln("center"),br=brd(),nf="+0.00%")
    cel(ws2,R2,5,sug,
        fl=fill(C4 if delta>0.005 else (C5 if delta<-0.005 else bg)),
        fn=fnt(b=True,c=sug_c,s=10),al=aln("center"),br=brd())
    ws2.row_dimensions[R2].height = 18; R2 += 1

ws2.merge_cells(f"A{R2+1}:E{R2+1}")
cel(ws2,R2+1,1,
    f"Otimizador Max Sharpe (SLSQP)  |  α={ALPHA} (mínimo)  β={BETA} (máximo)  "
    f"|  RF={RF_AA:.2%}  |  Status: {'OK' if ok else 'Convergência parcial'}",
    fl=fill(C3),fn=fnt(c=CO,s=9,i=True),al=aln("left"),br=brd())

# ─────────────────────────────────────────────
# ABA 3: DISCLAIMERS
# ─────────────────────────────────────────────
ws3 = wb.create_sheet("Disclaimers")
ws3.column_dimensions["A"].width = 90
disclaimers = [
    ("DISCLAIMERS E LIMITAÇÕES METODOLÓGICAS", True),
    ("", False),
    ("1. Esta análise é exclusivamente retroativa (backtesting). Os resultados refletem o comportamento "
     "passado dos ativos no período analisado e NÃO constituem previsão de desempenho futuro.", False),
    ("", False),
    ("2. NÃO é recomendação de investimento, análise de valores mobiliários nem consultoria financeira. "
     "Desenvolvido para fins educacionais e de pesquisa em finanças quantitativas.", False),
    ("", False),
    ("3. Os pesos da carteira são calculados com base na última cotação disponível no CSV de preços. "
     "Preços reais de mercado podem diferir.", False),
    ("", False),
    ("4. O otimizador maximiza o Índice de Sharpe histórico. Em janelas curtas, o modelo tende a "
     "superestimar o Sharpe in-sample (estimation error maximization). Resultados out-of-sample são "
     "tipicamente inferiores.", False),
    ("", False),
    ("5. Os indicadores NÃO incorporam dividendos, custos de transação, imposto de renda, "
     "spreads bid-ask ou outros custos operacionais.", False),
    ("", False),
    ("6. DIVO11 é utilizado como proxy do IDIV. Não é idêntico: há come-cotas, taxa de administração "
     "e tracking error.", False),
    ("", False),
    ("7. Antes de tomar qualquer decisão de investimento, consulte um analista credenciado pela CVM "
     "ou consultor de investimentos registrado.", False),
    ("", False),
    ("Metodologia completa e código-fonte:", False),
    ("  github.com/alissondpoliveira/portfolio-dividendos-b3", False),
]
for i, (line, bold) in enumerate(disclaimers):
    ws3.row_dimensions[i+1].height = 16
    is_title = (i == 0)
    cel(ws3, i+1, 1, line,
        fl=fill(C1 if is_title else ("F8F8F8" if i%2==0 else "FFFFFF")),
        fn=fnt(b=(bold or is_title), c=(C2 if is_title else "1A1A2E"),
               s=(11 if is_title else 9)),
        al=aln("left", w=True))
    ws3.row_dimensions[i+1].height = 30 if is_title else 16

wb.save(SAIDA)

# ══════════════════════════════════════════════════════════════════════
print(f"\n{'='*62}")
print(f"  Resultado salvo em: {SAIDA}")
print(f"{'='*62}")
print(f"\n  {'Metrica':<28} {'Atual':>10}  {'Otimizado':>10}")
print(f"  {'-'*50}")
for label, key, fmt in [
    ("Retorno Anual",  "Ra",      ".1%"),
    ("Volatilidade",   "vol",     ".1%"),
    ("Sharpe",         "sharpe",  ".3f"),
    ("Sortino",        "sortino", ".3f"),
    ("MDD",            "mdd",     ".1%"),
    ("Beta vs IBOV",   "beta",    ".3f"),
    ("Alpha Jensen",   "alpha_j", ".2%"),
]:
    vr = m_rec.get(key); vo = m_opt.get(key)
    sr = format(vr, fmt) if vr is not None else "—"
    so = format(vo, fmt) if vo is not None else "—"
    print(f"  {label:<28} {sr:>10}  {so:>10}")

print(f"\n  Valor estimado da carteira: R$ {valor_total:,.2f}")
print(f"\n  ⚠ Analise retroativa. Nao e recomendacao de investimento.")
